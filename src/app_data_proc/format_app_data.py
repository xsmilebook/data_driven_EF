#!/usr/bin/env python3
"""
Rename Excel sheet names for app behavioral data to consistent task names.
"""

import argparse
import glob
import os
import re
from pathlib import Path

import openpyxl

from src.config_io import load_simple_yaml
from src.path_config import load_paths_config, resolve_dataset_roots

def get_task_renaming_rules():
    """Return explicit sheet name renaming rules."""
    return {
        "LG": "EmotionSwitch",
        "STROOP": "ColorStroop",
        "SpatialNBack": "Spatial2Back",
        "Emotion2Backformal": "Emotion2Back",
        "Number2Backformal": "Number2Back",
        "Spatial1Backformal": "Spatial1Back",
        "EmotionStoop": "EmotionStroop",
    }


def normalize_sheet_name(sheet_name):
    """Normalize sheet name by applying explicit mapping and stripping 'formal'."""
    renaming_rules = get_task_renaming_rules()
    if sheet_name in renaming_rules:
        return renaming_rules[sheet_name]

    if "formal" in sheet_name.lower():
        return re.sub("formal", "", sheet_name, flags=re.IGNORECASE).strip()

    return sheet_name


def rename_sheets_in_excel(file_path, dry_run=False):
    """
    Rename sheets in a single Excel file.

    Returns:
        tuple[list[str], list[str]]: renamed sheets, skipped renames due to conflicts.
    """
    renamed_sheets = []
    skipped_conflicts = []

    try:
        workbook = openpyxl.load_workbook(file_path)
        original_names = list(workbook.sheetnames)

        for old_name in original_names:
            new_name = normalize_sheet_name(old_name)
            if new_name == old_name:
                continue
            if new_name in workbook.sheetnames:
                skipped_conflicts.append(f"'{old_name}' -> '{new_name}' (conflict)")
                continue

            workbook[old_name].title = new_name
            renamed_sheets.append(f"'{old_name}' -> '{new_name}'")

        if renamed_sheets and not dry_run:
            workbook.save(file_path)
        workbook.close()

        return renamed_sheets, skipped_conflicts
    except Exception as exc:
        print(f"Error processing {file_path}: {exc}")
        return [], []


def _get_header_col(ws, header_name):
    if ws is None:
        return None
    for cell in ws[1]:
        if str(cell.value).strip() == header_name:
            return cell.column
    return None


def _column_values(ws, col_idx):
    if ws is None or col_idx is None:
        return []
    return [ws.cell(row=r, column=col_idx).value for r in range(1, ws.max_row + 1)]


def _first_numeric_value(ws, col_idx):
    if ws is None or col_idx is None:
        return None
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=col_idx).value
        if val is None or str(val).strip() == '':
            continue
        try:
            return float(val)
        except Exception:
            continue
    return None


def _apply_reference_fix(workbook, reference_data):
    changes = []

    # SST: SSRT first numeric value is 0
    sst_ws = workbook['SST'] if 'SST' in workbook.sheetnames else None
    if sst_ws is not None:
        ssrt_col = _get_header_col(sst_ws, 'SSRT')
        first_num = _first_numeric_value(sst_ws, ssrt_col)
        if first_num == 0:
            ref_vals = reference_data.get('SST_SSRT', [])
            if ssrt_col is not None and ref_vals:
                for r in range(2, sst_ws.max_row + 1):
                    if r - 1 < len(ref_vals):
                        sst_ws.cell(row=r, column=ssrt_col).value = ref_vals[r - 1]
                    else:
                        sst_ws.cell(row=r, column=ssrt_col).value = None
                changes.append('SST SSRT column replaced')

    # Emotion1Back/Emotion2Back: first item is EmotionXBack_7
    item_header = '正式阶段刺激图片/Item名'
    for sheet_name, ref_key, first_val in [
        ('Emotion1Back', 'Emotion1Back_item', 'Emotion1Back_4'),
        ('Emotion2Back', 'Emotion2Back_item', 'Emotion2Back_7'),
    ]:
        ws = workbook[sheet_name] if sheet_name in workbook.sheetnames else None
        if ws is None:
            continue
        item_col = _get_header_col(ws, item_header)
        if item_col is None:
            continue
        first_cell = ws.cell(row=2, column=item_col).value
        if str(first_cell).strip() != first_val:
            continue
        ref_vals = reference_data.get(ref_key, [])
        if not ref_vals:
            continue
        for r in range(2, ws.max_row + 1):
            if r - 1 < len(ref_vals):
                ws.cell(row=r, column=item_col).value = ref_vals[r - 1]
            else:
                ws.cell(row=r, column=item_col).value = None
        changes.append(f'{sheet_name} item column replaced')

    return changes


def _load_reference_data(reference_path):
    data = {}
    if not os.path.exists(reference_path):
        return data
    wb = openpyxl.load_workbook(reference_path, data_only=True)
    try:
        sst_ws = wb['SST'] if 'SST' in wb.sheetnames else None
        if sst_ws is not None:
            ssrt_col = _get_header_col(sst_ws, 'SSRT')
            data['SST_SSRT'] = _column_values(sst_ws, ssrt_col)

        item_header = '正式阶段刺激图片/Item名'
        for sheet_name, key in [('Emotion1Back', 'Emotion1Back_item'),
                                ('Emotion2Back', 'Emotion2Back_item')]:
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else None
            if ws is None:
                continue
            item_col = _get_header_col(ws, item_header)
            data[key] = _column_values(ws, item_col)
    finally:
        wb.close()
    return data


def process_directory(directory_path, pattern="*.xlsx", dry_run=False):
    """Process all Excel files in the given directory."""
    excel_files = glob.glob(os.path.join(directory_path, pattern))
    if not excel_files:
        print(f"No Excel files found in {directory_path} with pattern {pattern}")
        return

    print(f"Found {len(excel_files)} Excel files to process")

    reference_name = 'THU_20250829_759_JZC_景志成_GameData.xlsx'
    reference_path = os.path.join(directory_path, reference_name)
    reference_data = _load_reference_data(reference_path)
    if not reference_data:
        print(f"Reference file not found or empty: {reference_path}")

    total_renamed_files = 0
    total_renamed_sheets = 0
    total_conflicts = 0
    total_fixed_files = 0
    total_fix_actions = 0

    for idx, file_path in enumerate(excel_files, start=1):
        print(f"\nProcessing file {idx}/{len(excel_files)}: {os.path.basename(file_path)}")
        renamed_sheets, conflicts = rename_sheets_in_excel(file_path, dry_run=dry_run)

        if renamed_sheets:
            total_renamed_files += 1
            total_renamed_sheets += len(renamed_sheets)
            print(f"  Renamed {len(renamed_sheets)} sheets: {', '.join(renamed_sheets)}")

        if conflicts:
            total_conflicts += len(conflicts)
            print(f"  Skipped {len(conflicts)} due to name conflicts: {', '.join(conflicts)}")

        if not reference_data or os.path.basename(file_path) == reference_name:
            continue

        try:
            workbook = openpyxl.load_workbook(file_path)
            changes = _apply_reference_fix(workbook, reference_data)
            if changes:
                total_fixed_files += 1
                total_fix_actions += len(changes)
                print(f"  Applied {len(changes)} data fixes: {', '.join(changes)}")
                if not dry_run:
                    workbook.save(file_path)
            workbook.close()
        except Exception as exc:
            print(f"  Data fix error in {file_path}: {exc}")

    print("\nProcessing complete!")
    print(f"Total files processed: {len(excel_files)}")
    print(f"Files with renamed sheets: {total_renamed_files}")
    print(f"Total sheets renamed: {total_renamed_sheets}")
    if total_fix_actions:
        print(f"Files with data fixes: {total_fixed_files}")
        print(f"Total data fix actions: {total_fix_actions}")
    if total_conflicts:
        print(f"Total rename conflicts: {total_conflicts}")


def _resolve_defaults(args):
    if not args.dataset:
        raise ValueError("Missing --dataset (required when defaults are used).")
    repo_root = Path(__file__).resolve().parents[2]
    paths_cfg = load_paths_config(args.paths_config, repo_root=repo_root)
    roots = resolve_dataset_roots(paths_cfg, dataset=args.dataset)
    dataset_cfg_path = (
        Path(args.dataset_config)
        if args.dataset_config is not None
        else (repo_root / "configs" / "datasets" / f"{args.dataset}.yaml")
    )
    dataset_cfg = load_simple_yaml(dataset_cfg_path)
    behavioral_cfg = dataset_cfg.get("behavioral", {})
    app_data_rel = behavioral_cfg.get("app_data_dir")
    if not app_data_rel:
        raise ValueError("Missing behavioral.app_data_dir in dataset config.")
    input_dir = roots["raw_root"] / app_data_rel
    return input_dir


def main():
    parser = argparse.ArgumentParser(
        description="Normalize Excel sheet names for app behavioral data."
    )
    parser.add_argument(
        "--input_dir",
        default=None,
        help="Directory containing Excel files.",
    )
    parser.add_argument(
        "--pattern",
        default="*.xlsx",
        help="Glob pattern for Excel files.",
    )
    parser.add_argument(
        "--dry_run",
        action="store_true",
        help="Print changes without saving files.",
    )
    parser.add_argument("--dataset", type=str, default=None)
    parser.add_argument("--config", dest="paths_config", type=str, default="configs/paths.yaml")
    parser.add_argument("--dataset-config", dest="dataset_config", type=str, default=None)

    args = parser.parse_args()

    if args.input_dir is None:
        input_dir = _resolve_defaults(args)
        args.input_dir = str(input_dir)

    if not os.path.exists(args.input_dir):
        print(f"Error: Directory does not exist: {args.input_dir}")
        return
    if not os.path.isdir(args.input_dir):
        print(f"Error: Path is not a directory: {args.input_dir}")
        return

    print(f"Starting to process Excel files in: {args.input_dir}")
    process_directory(args.input_dir, pattern=args.pattern, dry_run=args.dry_run)
    print("Script completed successfully!")


if __name__ == "__main__":
    main()
