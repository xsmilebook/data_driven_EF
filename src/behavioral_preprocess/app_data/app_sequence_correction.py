from __future__ import annotations

import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import numpy as np
import openpyxl
import pandas as pd


ITEM_COL = "正式阶段刺激图片/Item名"
ANSWER_COL = "正式阶段正确答案"
BLANK_SCREEN_COL = "空屏时长"


def parse_subject_id_from_filename(excel_path: Path) -> str:
    stem = excel_path.stem
    stem = stem.replace("GameData", "").rstrip("_- ")
    parts = [p for p in stem.split("_") if p]
    if len(parts) >= 4:
        return "_".join(parts[:4])
    return stem


def parse_subject_date(subject_id: str) -> date | None:
    # Expected: THU_YYYYMMDD_...
    parts = subject_id.split("_")
    if len(parts) < 2:
        return None
    s = parts[1]
    if not re.fullmatch(r"\d{8}", s):
        return None
    try:
        return datetime.strptime(s, "%Y%m%d").date()
    except Exception:
        return None


def expected_visit_by_date(d: date | None) -> str | None:
    if d is None:
        return None
    if d < date(2025, 7, 8):
        return "visit1"
    if d <= date(2025, 11, 27):
        return "visit3"
    if d <= date(2025, 12, 15):
        return "visit2"
    if d <= date(2026, 1, 12):
        return "visit1"
    return "visit4"


def _normalize_header(value: object) -> str:
    s = str(value).strip()
    return "".join(s.split())


def _normalize_cell(value: object) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    low = s.strip().lower()
    if low in {"nan", "none"}:
        return None
    s = s.replace("\\", "/")
    if "/" in s:
        s = s.split("/")[-1]
    s = s.strip().lower()
    for ext in (".png", ".jpg", ".jpeg", ".bmp", ".gif", ".tif", ".tiff", ".webp"):
        if s.endswith(ext):
            s = s[: -len(ext)]
            break
    if s.isdigit():
        s2 = s.lstrip("0")
        s = s2 if s2 else "0"
    return s or None


def _canonical_task_name(raw: str) -> str:
    s = str(raw).strip()
    if not s:
        return s
    rules = {
        "LG": "EmotionSwitch",
        "STROOP": "ColorStroop",
        "SpatialNBack": "Spatial2Back",
        "Emotion2Backformal": "Emotion2Back",
        "Number2Backformal": "Number2Back",
        "Spatial1Backformal": "Spatial1Back",
        "EmotionStoop": "EmotionStroop",
    }
    if s in rules:
        return rules[s]
    if "formal" in s.lower():
        s = re.sub("formal", "", s, flags=re.IGNORECASE).strip()
    return s


def _find_col_by_header(df: pd.DataFrame, target: str, *, fallback_contains: list[str] | None = None) -> str | None:
    target_norm = _normalize_header(target)
    cols = [str(c) for c in df.columns]
    for c in cols:
        if _normalize_header(c) == target_norm:
            return c
    if fallback_contains:
        for c in cols:
            cn_low = _normalize_header(c).lower()
            if all(str(tok).lower() in cn_low for tok in fallback_contains):
                return c
    return None


def _clean_raw_value(value: object) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    low = s.lower()
    if low in {"nan", "none"}:
        return None
    s = s.replace("\\", "/")
    if "/" in s:
        s = s.split("/")[-1]
    return s.strip() or None


@dataclass(frozen=True)
class TaskSequence:
    task: str
    items_raw: list[str | None]
    answers_raw: list[str | None]
    items_norm: list[str | None]
    answers_norm: list[str | None]
    source: str


@dataclass(frozen=True)
class ExportSourceDecision:
    export_source: str
    relevant_sheets: int
    sheets_with_blank_screen_col: int
    blank_screen_in_all_sheets: bool
    sheets_missing_blank_screen_col: list[str]


def infer_export_source_by_blank_screen(excel_path: Path) -> ExportSourceDecision:
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    relevant: list[str] = []
    with_col: list[str] = []
    missing: list[str] = []
    target_norm = _normalize_header(BLANK_SCREEN_COL)
    item_norm = _normalize_header(ITEM_COL)
    answer_norm = _normalize_header(ANSWER_COL)
    task_norm = _normalize_header("任务")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration:
            continue
        header_norms = {_normalize_header(v) for v in header if v is not None}
        if not header_norms:
            continue
        is_relevant = (item_norm in header_norms) or (answer_norm in header_norms) or (task_norm in header_norms)
        if not is_relevant:
            continue
        relevant.append(sheet_name)
        if target_norm in header_norms:
            with_col.append(sheet_name)
        else:
            missing.append(sheet_name)

    wb.close()

    if not relevant:
        return ExportSourceDecision(
            export_source="unknown",
            relevant_sheets=0,
            sheets_with_blank_screen_col=0,
            blank_screen_in_all_sheets=False,
            sheets_missing_blank_screen_col=[],
        )

    blank_in_all = len(missing) == 0
    export_source = "txt_export" if blank_in_all else "web_export"
    return ExportSourceDecision(
        export_source=export_source,
        relevant_sheets=len(relevant),
        sheets_with_blank_screen_col=len(with_col),
        blank_screen_in_all_sheets=blank_in_all,
        sheets_missing_blank_screen_col=missing,
    )


def _detect_items_key(obj: dict[str, Any]) -> str | None:
    keys = [k for k in obj.keys() if isinstance(k, str)]
    for k in keys:
        if k.lower().endswith("_items"):
            return k
    return None


def _task_from_items_key(key: str) -> str:
    s = key
    if s.lower().endswith("_items"):
        s = s[: -len("_Items")]
    # Remove trailing "Formal" and optional digits (e.g., Formal20251125).
    s = re.sub(r"formal\d*$", "", s, flags=re.IGNORECASE)
    s = s.strip("_- ")
    return _canonical_task_name(s)


def _task_from_filename(path: Path) -> str:
    stem = path.stem
    stem = re.sub(r"formal\d*$", "", stem, flags=re.IGNORECASE)
    stem = re.sub(r"\d{4,8}$", "", stem)  # e.g., stroop1128
    stem = stem.strip("_- ")
    return _canonical_task_name(stem)


def _extract_item_name(record: dict[str, Any], *, task: str) -> str | None:
    # Prefer explicit PicName-like keys
    for k in (
        "picData",
        "PicData",
        "step2_PicName",
        "step1_PicName",
        "PicName",
        "picName",
        "step2_picName",
    ):
        if k in record and record[k] is not None:
            return _clean_raw_value(record[k])

    # SST: use isLeftPic + isChangePic when no PicName exists.
    if "isLeftPic" in record:
        side = "Left" if bool(record.get("isLeftPic")) else "Right"
        if bool(record.get("isChangePic")):
            return f"{side}_Stop"
        return side

    return None


def _extract_answer_name(record: dict[str, Any]) -> str | None:
    if "buttonName" in record:
        return _clean_raw_value(record.get("buttonName"))
    if "ButtonName" in record:
        return _clean_raw_value(record.get("ButtonName"))
    return None


def load_visit_sequences(visit_dir: Path) -> dict[str, TaskSequence]:
    out: dict[str, TaskSequence] = {}
    for p in sorted(visit_dir.glob("*")):
        if p.is_dir():
            continue
        if p.suffix.lower() not in {".json", ".txt"}:
            continue
        try:
            raw = p.read_text(encoding="utf-8")
            obj = json.loads(raw)
        except Exception:
            # Some txt/json may not be UTF-8; try system default as fallback.
            try:
                raw = p.read_text(encoding="gbk")
                obj = json.loads(raw)
            except Exception:
                continue

        if not isinstance(obj, dict):
            continue
        items_key = _detect_items_key(obj)
        if not items_key:
            continue
        items_raw = obj.get(items_key)
        if not isinstance(items_raw, list):
            continue

        task = _task_from_items_key(items_key) or _task_from_filename(p)
        items_out: list[str | None] = []
        answers_out: list[str | None] = []
        for rec in items_raw:
            if not isinstance(rec, dict):
                items_out.append(None)
                answers_out.append(None)
                continue
            items_out.append(_extract_item_name(rec, task=task))
            answers_out.append(_extract_answer_name(rec))

        items_norm = [_normalize_cell(x) for x in items_out]
        answers_norm = [_normalize_cell(x) for x in answers_out]
        out[task] = TaskSequence(
            task=task,
            items_raw=items_out,
            answers_raw=answers_out,
            items_norm=items_norm,
            answers_norm=answers_norm,
            source=str(p),
        )
    return out


def load_all_visits_sequences(sequence_root: Path) -> dict[str, dict[str, TaskSequence]]:
    """
    Load visit sequence configs from app_sequence root.

    Expected directories:
      - visit1/
      - visit2*/ (e.g., visit2-1125)
      - visit3/
      - visit4/
    """
    visits: dict[str, dict[str, TaskSequence]] = {}
    for d in sorted(sequence_root.iterdir()):
        if not d.is_dir():
            continue
        name = d.name
        visit = None
        if name == "visit1":
            visit = "visit1"
        elif name.startswith("visit2"):
            visit = "visit2"
        elif name == "visit3":
            visit = "visit3"
        elif name == "visit4":
            visit = "visit4"
        if not visit:
            continue
        visits[visit] = load_visit_sequences(d)
    return visits


def build_effective_visit_sequences(
    visits: dict[str, dict[str, TaskSequence]],
    *,
    baseline_visit: str = "visit1",
) -> dict[str, dict[str, TaskSequence]]:
    """
    Fill missing tasks in each visit using baseline visit (default visit1),
    since visit2~4 configs may be incomplete.
    """
    if baseline_visit not in visits:
        raise ValueError(f"Missing baseline visit: {baseline_visit}")
    baseline = visits[baseline_visit]
    out: dict[str, dict[str, TaskSequence]] = {}
    for visit, seqs in visits.items():
        merged: dict[str, TaskSequence] = dict(baseline)
        merged.update(seqs)
        out[visit] = merged
    return out


@dataclass(frozen=True)
class ObservedTaskData:
    task: str
    n_rows: int
    item_seq: list[str | None] | None
    answer_seq: list[str | None] | None


def extract_observed_task_data(excel_path: Path) -> dict[str, ObservedTaskData]:
    xl = pd.ExcelFile(excel_path)
    out: dict[str, ObservedTaskData] = {}
    for sheet in xl.sheet_names:
        task = _canonical_task_name(sheet)
        df = xl.parse(sheet, dtype="object")
        if "任务" in df.columns and len(df) == 97:
            try:
                t0 = str(df["任务"].iloc[0]).strip()
            except Exception:
                t0 = ""
            if t0.upper() == "SST":
                df = df.iloc[:-1].copy()

        item_col = _find_col_by_header(
            df,
            ITEM_COL,
            fallback_contains=["正式阶段刺激图片", "Item"],
        )
        ans_col = _find_col_by_header(
            df,
            ANSWER_COL,
            fallback_contains=["正式阶段正确答案"],
        )
        item_seq = None
        answer_seq = None
        if item_col is not None:
            item_seq = [_normalize_cell(x) for x in df[item_col].tolist()]
        if ans_col is not None:
            answer_seq = [_normalize_cell(x) for x in df[ans_col].tolist()]
        out[task] = ObservedTaskData(
            task=task,
            n_rows=int(len(df)),
            item_seq=item_seq,
            answer_seq=answer_seq,
        )
    return out


def _seq_match_score(obs: list[str | None], exp: list[str | None]) -> float:
    if not obs or not exp:
        return float("nan")
    n = min(len(obs), len(exp))
    if n == 0:
        return float("nan")
    matches = 0
    for a, b in zip(obs[:n], exp[:n], strict=False):
        if a == b:
            matches += 1
    base = matches / n
    # Penalize length mismatch mildly.
    max_len = max(len(obs), len(exp))
    penalty = abs(len(obs) - len(exp)) / max_len if max_len else 0.0
    return float(base - 0.25 * penalty)


@dataclass(frozen=True)
class SubjectVisitInference:
    subject_id: str
    excel_path: Path
    inferred_visit: str
    score: float
    expected_visit: str | None
    scores_by_visit: dict[str, float]


def infer_subject_visit(
    excel_path: Path,
    subject_id: str,
    observed: dict[str, ObservedTaskData],
    visits: dict[str, dict[str, TaskSequence]],
) -> SubjectVisitInference:
    d = parse_subject_date(subject_id)
    expected = expected_visit_by_date(d)
    scores: dict[str, float] = {}
    for visit, seqs in visits.items():
        task_scores: list[float] = []
        for task, obs in observed.items():
            if obs.answer_seq is None:
                continue
            exp = seqs.get(task)
            if exp is None:
                continue
            sc = _seq_match_score(obs.answer_seq, exp.answers_norm)
            if np.isnan(sc):
                continue
            task_scores.append(sc)
        if task_scores:
            mean = float(np.mean(task_scores))
        else:
            mean = float("nan")
        # Small prior boost if matches expected visit by date.
        if expected and visit == expected and not np.isnan(mean):
            mean += 0.02
        scores[visit] = mean

    best_visit = None
    best_score = -1e9
    for visit, sc in scores.items():
        if np.isnan(sc):
            continue
        if sc > best_score:
            best_score = sc
            best_visit = visit
    if best_visit is None:
        # Fallback: expected by date, else visit1.
        best_visit = expected or "visit1"
        best_score = float("nan")
    return SubjectVisitInference(
        subject_id=subject_id,
        excel_path=excel_path,
        inferred_visit=best_visit,
        score=best_score,
        expected_visit=expected,
        scores_by_visit=scores,
    )


@dataclass(frozen=True)
class TaskCorrectionDecision:
    task: str
    visit: str
    item_match: float | None
    answer_match: float | None
    issue_type: str | None


@dataclass(frozen=True)
class WorkbookCorrectionResult:
    subject_id: str
    inferred_visit: str
    output_excel: Path
    task_decisions: list[TaskCorrectionDecision]
    note: str | None = None


def _classify_issue(
    *,
    item_match_to_visit: float | None,
    answer_match_to_visit: float | None,
    item_match_best_other: float | None,
    answer_match_best_other: float | None,
) -> str | None:
    # Heuristic labels aligned with the described export issues.
    if answer_match_to_visit is not None and answer_match_to_visit >= 0.95:
        if item_match_to_visit is not None and item_match_to_visit < 0.8:
            if item_match_best_other is not None and item_match_best_other >= 0.95:
                return "backend_item_overwritten_suspected"
            return "item_mismatch_suspected"
    if answer_match_to_visit is not None and answer_match_to_visit < 0.8:
        if answer_match_best_other is not None and answer_match_best_other >= 0.95:
            return "answer_version_inconsistent"
    return None


def correct_workbook(
    *,
    excel_path: Path,
    output_path: Path,
    inferred_visit: str,
    visits: dict[str, dict[str, TaskSequence]],
) -> WorkbookCorrectionResult:
    subject_id = parse_subject_id_from_filename(excel_path)
    observed = extract_observed_task_data(excel_path)
    visit_seqs = visits[inferred_visit]

    wb = openpyxl.load_workbook(excel_path)
    task_decisions: list[TaskCorrectionDecision] = []

    for sheet_name in wb.sheetnames:
        task = _canonical_task_name(sheet_name)
        seq = visit_seqs.get(task)
        if seq is None:
            continue

        ws = wb[sheet_name]
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        item_col_idx = None
        ans_col_idx = None
        for idx, val in enumerate(header, start=1):
            if val is None:
                continue
            if _normalize_header(val) == _normalize_header(ITEM_COL):
                item_col_idx = idx
            if _normalize_header(val) == _normalize_header(ANSWER_COL):
                ans_col_idx = idx

        if task.upper() == "SST" and ws.max_row == 98:
            # 1 header + 97 data rows -> drop last invalid row
            ws.delete_rows(ws.max_row)

        n_rows = ws.max_row - 1
        exp_items_norm = seq.items_norm
        exp_answers_norm = seq.answers_norm
        exp_items_raw = seq.items_raw
        exp_answers_raw = seq.answers_raw
        # Match scores (for manifest/debugging)
        obs_task = observed.get(task)
        item_match = None
        ans_match = None
        if obs_task and obs_task.item_seq is not None:
            item_match = _seq_match_score(obs_task.item_seq, exp_items_norm)
        if obs_task and obs_task.answer_seq is not None:
            ans_match = _seq_match_score(obs_task.answer_seq, exp_answers_norm)

        best_other_item = None
        best_other_ans = None
        for v, seqs in visits.items():
            if v == inferred_visit:
                continue
            other = seqs.get(task)
            if other is None or obs_task is None:
                continue
            if obs_task.item_seq is not None:
                sc = _seq_match_score(obs_task.item_seq, other.items_norm)
                if not np.isnan(sc):
                    best_other_item = sc if best_other_item is None else max(best_other_item, sc)
            if obs_task.answer_seq is not None:
                sc = _seq_match_score(obs_task.answer_seq, other.answers_norm)
                if not np.isnan(sc):
                    best_other_ans = sc if best_other_ans is None else max(best_other_ans, sc)

        issue_type = _classify_issue(
            item_match_to_visit=item_match,
            answer_match_to_visit=ans_match,
            item_match_best_other=best_other_item,
            answer_match_best_other=best_other_ans,
        )
        task_decisions.append(
            TaskCorrectionDecision(
                task=task,
                visit=inferred_visit,
                item_match=item_match,
                answer_match=ans_match,
                issue_type=issue_type,
            )
        )

        # Apply corrected sequences to the workbook in-place (but write to output_path).
        if item_col_idx is not None:
            for i in range(1, n_rows + 1):
                val = exp_items_raw[i - 1] if i - 1 < len(exp_items_raw) else None
                ws.cell(row=i + 1, column=item_col_idx).value = val
        if ans_col_idx is not None:
            for i in range(1, n_rows + 1):
                val = exp_answers_raw[i - 1] if i - 1 < len(exp_answers_raw) else None
                ws.cell(row=i + 1, column=ans_col_idx).value = val

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()
    return WorkbookCorrectionResult(
        subject_id=subject_id,
        inferred_visit=inferred_visit,
        output_excel=output_path,
        task_decisions=task_decisions,
    )
