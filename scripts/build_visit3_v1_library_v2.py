from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

import openpyxl

from src.path_config import load_dataset_config, load_paths_config, resolve_dataset_roots


TARGET_HEADER = "\u6b63\u5f0f\u9636\u6bb5\u6b63\u786e\u7b54\u6848"
TARGET_SUBKEY = "\u6b63\u786e\u7b54\u6848"


def _normalize_header(value: object) -> str:
    s = str(value).strip()
    return "".join(s.split())


def _canonical_task_name(raw: str) -> str:
    s = str(raw).strip()
    if not s:
        return s
    s = s.strip("_- ")
    rules = {
        "EmotionStoop": "EmotionStroop",
        "LG": "EmotionSwitch",
        "STROOP": "ColorStroop",
        "SpatialNBack": "Spatial2Back",
        "Emotion2Backformal": "Emotion2Back",
        "Number2Backformal": "Number2Back",
        "Spatial1Backformal": "Spatial1Back",
    }
    if s in rules:
        return rules[s]
    if "formal" in s.lower():
        s = re.sub("formal", "", s, flags=re.IGNORECASE).strip()
    return s


def _normalize_cell(value: object) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    low = s.lower()
    if low in {"nan", "none"}:
        return None
    s = s.replace("\\", "/")
    if "/" in s:
        s = s.split("/")[-1]
    s = s.strip().lower()
    for ext in (".png", ".jpg", ".jpeg", ".bmp", ".gif", ".tif", ".tiff", ".webp"):
        if s.endswith(ext):
            s = s[: -len(ext)]
            break
    if s.isdigit():
        s2 = s.lstrip("0")
        s = s2 if s2 else "0"
    return s or None


def _extract_subject_id_from_filename(excel_path: Path) -> str:
    stem = excel_path.stem
    stem = stem.replace("GameData", "").rstrip("_- ")
    parts = [p for p in stem.split("_") if p]
    if len(parts) >= 4:
        return "_".join(parts[:4])
    return stem


def _index_raw_workbooks(raw_app_dir: Path) -> dict[str, Path]:
    out: dict[str, Path] = {}
    for fp in raw_app_dir.glob("*.xlsx"):
        if fp.name.startswith("~$"):
            continue
        sid = _extract_subject_id_from_filename(fp)
        out.setdefault(sid, fp)
    return out


def _read_group_subjects(path: Path) -> list[str]:
    if not path.exists():
        raise FileNotFoundError(f"Missing group file: {path}")
    return [s.strip() for s in path.read_text(encoding="utf-8").splitlines() if s.strip()]


def _read_subject_task_answers(xlsx: Path, task: str) -> list[str]:
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    try:
        ws_name = None
        for n in wb.sheetnames:
            if _canonical_task_name(n) == task:
                ws_name = n
                break
        if ws_name is None:
            return []
        ws = wb[ws_name]
        header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
        ans_idx = None
        for i, h in enumerate(header, start=1):
            hv = _normalize_header(h)
            if hv == TARGET_HEADER or TARGET_SUBKEY in hv:
                ans_idx = i
                break
        if ans_idx is None:
            return []
        out: list[str] = []
        for r in range(2, ws.max_row + 1):
            v = _normalize_cell(ws.cell(r, ans_idx).value)
            if v is not None:
                out.append(v)
        return out
    finally:
        wb.close()


def _pick_group4_emotion2back_source(
    *,
    group4_subjects: list[str],
    raw_index: dict[str, Path],
) -> tuple[str, str, list[str]]:
    best_sid = ""
    best_file = ""
    best_answers: list[str] = []
    for sid in group4_subjects:
        fp = raw_index.get(sid)
        if fp is None:
            continue
        answers = _read_subject_task_answers(fp, "Emotion2Back")
        if len(answers) > len(best_answers):
            best_sid = sid
            best_file = fp.name
            best_answers = answers
    if not best_answers:
        raise ValueError("No non-empty Emotion2Back answers found in answer_group_004.")
    return best_sid, best_file, best_answers


def _load_visit3_v1_task_json(seq_dir: Path, filename: str, key: str) -> tuple[list[Any], list[Any]]:
    path = seq_dir / filename
    if not path.exists():
        raise FileNotFoundError(f"Missing visit3_v1 file: {path}")
    obj = json.loads(path.read_text(encoding="utf-8"))
    arr = obj.get(key)
    if not isinstance(arr, list):
        raise ValueError(f"Invalid structure in {path}: missing list key {key}")
    items: list[Any] = []
    answers: list[Any] = []
    for rec in arr:
        if not isinstance(rec, dict):
            continue
        items.append(rec.get("picData"))
        answers.append(rec.get("buttonName"))
    return items, answers


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(
        description=(
            "Build visit3_v1 in run_corrected_v2: use visit3 as baseline; replace Spatial1Back/Spatial2Back "
            "from app_sequence/visit3_v1, and set Emotion2Back items to empty with answers from answer_group_004."
        )
    )
    ap.add_argument("--dataset", type=str, required=True)
    ap.add_argument("--config", dest="paths_config", type=str, default="configs/paths.yaml")
    ap.add_argument("--dataset-config", dest="dataset_config", type=str, default=None)
    return ap.parse_args()


def main() -> int:
    args = parse_args()
    repo_root = Path(__file__).resolve().parents[1]
    paths_cfg = load_paths_config(args.paths_config, repo_root=repo_root)
    roots = resolve_dataset_roots(paths_cfg, dataset=args.dataset)
    dataset_cfg = load_dataset_config(paths_cfg, dataset_config_path=args.dataset_config, repo_root=repo_root)

    behavioral_cfg = dataset_cfg.get("behavioral", {})
    app_data_rel = behavioral_cfg.get("app_data_dir")
    app_seq_rel = behavioral_cfg.get("app_sequence_dir")
    corrected_rel = behavioral_cfg.get("corrected_app_excel_dir")
    interim_rel = behavioral_cfg.get("interim_preprocess_dir", "behavioral_preprocess")
    if not app_data_rel or not app_seq_rel or not corrected_rel:
        raise ValueError("Missing dataset.behavioral.app_data_dir/app_sequence_dir/corrected_app_excel_dir.")

    raw_app_dir = roots["raw_root"] / app_data_rel
    app_seq_root = roots["raw_root"] / app_seq_rel
    out_dir = roots["processed_root"] / corrected_rel / "run_corrected_v2" / "sequence_library"
    interim_dir = roots["interim_root"] / interim_rel

    visit3_base_path = out_dir / "visit3_from_groups.json"
    if not visit3_base_path.exists():
        raise FileNotFoundError(f"Missing baseline visit3 file: {visit3_base_path}")
    visit3 = json.loads(visit3_base_path.read_text(encoding="utf-8"))

    visit3_v1_dir = app_seq_root / "visit3_v1"
    sp1_items, sp1_answers = _load_visit3_v1_task_json(
        visit3_v1_dir,
        "spatial1back2.1.json",
        "Spatial1Back_Formal_Items",
    )
    sp2_items, sp2_answers = _load_visit3_v1_task_json(
        visit3_v1_dir,
        "spatial2back2.1.json",
        "Spatial2Back_Formal_Items",
    )

    raw_index = _index_raw_workbooks(raw_app_dir)
    group4_subjects = _read_group_subjects(interim_dir / "groups_by_answer" / "group_004_sublist.txt")
    e2_sid, e2_file, e2_answers = _pick_group4_emotion2back_source(
        group4_subjects=group4_subjects,
        raw_index=raw_index,
    )

    visit3_v1 = json.loads(json.dumps(visit3))

    # Override only the requested three tasks.
    visit3_v1["Spatial1Back"]["items"] = sp1_items
    visit3_v1["Spatial1Back"]["answers"] = sp1_answers
    visit3_v1["Spatial1Back"]["items_source"] = "app_sequence_visit3_v1"
    visit3_v1["Spatial1Back"]["answers_source"] = "app_sequence_visit3_v1"

    visit3_v1["Spatial2Back"]["items"] = sp2_items
    visit3_v1["Spatial2Back"]["answers"] = sp2_answers
    visit3_v1["Spatial2Back"]["items_source"] = "app_sequence_visit3_v1"
    visit3_v1["Spatial2Back"]["answers_source"] = "app_sequence_visit3_v1"

    visit3_v1["Emotion2Back"]["items"] = [None] * len(e2_answers)
    visit3_v1["Emotion2Back"]["answers"] = e2_answers
    visit3_v1["Emotion2Back"]["items_source"] = "forced_empty_by_rule"
    visit3_v1["Emotion2Back"]["answers_source"] = "answer_group_004_subject"
    visit3_v1["Emotion2Back"]["answer_template_subject_id"] = e2_sid
    visit3_v1["Emotion2Back"]["answer_template_file_name"] = e2_file

    out_json = out_dir / "visit3_v1_from_groups.json"
    out_meta = out_dir / "visit3_v1_build_meta.json"

    out_json.write_text(json.dumps(visit3_v1, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    out_meta.write_text(
        json.dumps(
            {
                "base": visit3_base_path.as_posix(),
                "rules": {
                    "Spatial1Back": "replace item/answer from app_sequence/visit3_v1/spatial1back2.1.json",
                    "Spatial2Back": "replace item/answer from app_sequence/visit3_v1/spatial2back2.1.json",
                    "Emotion2Back": "items forced to null; answers from answer_group_004",
                },
                "emotion2back_answer_source": {
                    "subject_id": e2_sid,
                    "file_name": e2_file,
                    "answers_len": len(e2_answers),
                },
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )

    print(f"out={out_json}")
    print(f"meta={out_meta}")
    print(f"emotion2back_source_subject={e2_sid}")
    print(f"emotion2back_answers_len={len(e2_answers)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

