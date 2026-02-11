from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

from src.path_config import load_dataset_config, load_paths_config, resolve_dataset_roots


ITEM_COL = "正式阶段刺激图片/Item名"
ANSWER_COL = "正式阶段正确答案"


def _normalize_header(value: object) -> str:
    s = str(value).strip()
    return "".join(s.split())


def _find_col_idx(header_row: list[object], target: str) -> int | None:
    target_norm = _normalize_header(target)
    for idx, val in enumerate(header_row, start=1):
        if val is None:
            continue
        if _normalize_header(val) == target_norm:
            return idx
    return None


def _read_column(ws: openpyxl.worksheet.worksheet.Worksheet, col_idx: int, *, max_row: int | None = None) -> list[Any]:
    mr = max_row or ws.max_row
    vals: list[Any] = []
    for row in ws.iter_rows(min_row=2, max_row=mr, min_col=col_idx, max_col=col_idx, values_only=True):
        vals.append(row[0])
    return vals


def _canonical_task_name(raw: str) -> str:
    s = str(raw).strip()
    if not s:
        return s
    s = s.strip("_- ")
    rules = {
        "EmotionStoop": "EmotionStroop",
        "LG": "EmotionSwitch",
        "STROOP": "ColorStroop",
        "SpatialNBack": "Spatial2Back",
        "Emotion2Backformal": "Emotion2Back",
        "Number2Backformal": "Number2Back",
        "Spatial1Backformal": "Spatial1Back",
    }
    if s in rules:
        return rules[s]
    if "formal" in s.lower():
        s = re.sub("formal", "", s, flags=re.IGNORECASE).strip()
    return s


def _extract_subject_id_from_filename(excel_path: Path) -> str:
    stem = excel_path.stem
    stem = stem.replace("GameData", "").rstrip("_- ")
    parts = [p for p in stem.split("_") if p]
    if len(parts) >= 4:
        return "_".join(parts[:4])
    return stem


def _normalize_cell(value: object) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    low = s.lower()
    if low in {"nan", "none"}:
        return None
    s = s.replace("\\", "/")
    if "/" in s:
        s = s.split("/")[-1]
    s = s.strip().lower()
    for ext in (".png", ".jpg", ".jpeg", ".bmp", ".gif", ".tif", ".tiff", ".webp"):
        if s.endswith(ext):
            s = s[: -len(ext)]
            break
    if s.isdigit():
        s2 = s.lstrip("0")
        s = s2 if s2 else "0"
    return s or None


def _normalize_seq_for_compare(seq: list[Any] | None) -> list[str]:
    if not seq:
        return []
    out: list[str] = []
    for x in seq:
        n = _normalize_cell(x)
        if n is not None:
            out.append(n)
    return out


def _match_ratio(a: list[str], b: list[str]) -> float | None:
    n = min(len(a), len(b))
    if n == 0:
        return None
    same = sum(1 for i in range(n) if a[i] == b[i])
    return same / n


@dataclass(frozen=True)
class TaskTemplate:
    task: str
    subject_id: str
    file_name: str
    n_rows: int
    sequence: list[Any] | None


def _index_raw_workbooks(raw_app_dir: Path) -> dict[str, Path]:
    out: dict[str, Path] = {}
    for fp in raw_app_dir.glob("*.xlsx"):
        if fp.name.startswith("~$"):
            continue
        sid = _extract_subject_id_from_filename(fp)
        out.setdefault(sid, fp)
    return out


def _read_subject_ids(path: Path) -> list[str]:
    if not path.exists():
        raise FileNotFoundError(f"Missing group file: {path}")
    return [s.strip() for s in path.read_text(encoding="utf-8").splitlines() if s.strip()]


def _build_templates_from_subjects(
    *,
    subject_ids: list[str],
    raw_index: dict[str, Path],
    target_col: str,
) -> dict[str, TaskTemplate]:
    templates: dict[str, TaskTemplate] = {}
    for sid in subject_ids:
        fp = raw_index.get(sid)
        if fp is None:
            continue
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        try:
            for sheet_name in wb.sheetnames:
                task = _canonical_task_name(sheet_name)
                if task in templates:
                    continue
                ws = wb[sheet_name]
                try:
                    header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
                except StopIteration:
                    continue
                col_idx = _find_col_idx(header, target_col)
                if col_idx is None:
                    continue
                max_row = ws.max_row
                if task.upper() == "SST" and max_row == 98:
                    max_row = 97
                seq = _read_column(ws, col_idx, max_row=max_row)
                templates[task] = TaskTemplate(
                    task=task,
                    subject_id=sid,
                    file_name=fp.name,
                    n_rows=max_row - 1,
                    sequence=seq,
                )
        finally:
            wb.close()
    return templates


def _read_app_sequence_visit(seq_root: Path, visit_name: str) -> dict[str, dict[str, Any]]:
    visit_dir = seq_root / visit_name
    out: dict[str, dict[str, Any]] = {}
    if not visit_dir.exists():
        return out
    for p in sorted(visit_dir.glob("*")):
        if p.is_dir() or p.suffix.lower() not in {".json", ".txt"}:
            continue
        try:
            obj = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            try:
                obj = json.loads(p.read_text(encoding="gbk"))
            except Exception:
                continue
        if not isinstance(obj, dict):
            continue
        items_key = next((k for k in obj.keys() if isinstance(k, str) and k.lower().endswith("_items")), None)
        if not items_key:
            continue
        raw_items = obj.get(items_key)
        if not isinstance(raw_items, list):
            continue
        task = re.sub(r"_items$", "", items_key, flags=re.IGNORECASE)
        task = re.sub(r"formal\d*$", "", task, flags=re.IGNORECASE)
        task = _canonical_task_name(task).strip("_- ")

        items: list[Any] = []
        answers: list[Any] = []
        for rec in raw_items:
            if not isinstance(rec, dict):
                continue

            item_val = None
            for k in ("step2_PicName", "step1_PicName", "PicName", "picName", "picData", "PicData"):
                if k in rec and rec[k] is not None:
                    item_val = rec[k]
                    break
            if item_val is not None:
                items.append(item_val)

            ans_val = None
            for k in ("buttonName", "ButtonName"):
                if k in rec and rec[k] is not None:
                    ans_val = rec[k]
                    break
            if ans_val is None and "needClickButton" in rec:
                ans_val = rec.get("needClickButton")
            if ans_val is None and isinstance(rec.get("answers"), list):
                for a in rec["answers"]:
                    if isinstance(a, dict) and a.get("answerPicName") is not None:
                        answers.append(a.get("answerPicName"))
            if ans_val is not None:
                answers.append(ans_val)

        out[task] = {
            "items": items,
            "answers": answers,
            "source": [p.as_posix()],
        }
    return out


def _compare_to_app_sequence(
    *,
    visit3_groups: dict[str, dict[str, Any]],
    visit3_app_seq: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    rows: dict[str, Any] = {}
    for task in sorted(set(visit3_groups) | set(visit3_app_seq)):
        g = visit3_groups.get(task, {})
        a = visit3_app_seq.get(task, {})

        g_items = _normalize_seq_for_compare(g.get("items"))
        g_answers = _normalize_seq_for_compare(g.get("answers"))
        a_items = _normalize_seq_for_compare(a.get("items"))
        a_answers = _normalize_seq_for_compare(a.get("answers"))

        item_ratio = _match_ratio(g_items, a_items)
        ans_ratio = _match_ratio(g_answers, a_answers)

        rows[task] = {
            "items_match_exact": bool(item_ratio == 1.0 and len(g_items) == len(a_items) and len(g_items) > 0),
            "answers_match_exact": bool(ans_ratio == 1.0 and len(g_answers) == len(a_answers) and len(g_answers) > 0),
            "items_match_ratio": item_ratio,
            "answers_match_ratio": ans_ratio,
            "groups_items_len": len(g_items),
            "groups_answers_len": len(g_answers),
            "app_items_len": len(a_items),
            "app_answers_len": len(a_answers),
            "app_sequence_source": a.get("source"),
        }
    return rows


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(
        description=(
            "Build visit3 sequence library in v2 using item_group_002 as item source "
            "and answer_group_002 as answer source, then compare with app_sequence/visit3."
        )
    )
    ap.add_argument("--dataset", type=str, required=True)
    ap.add_argument("--config", dest="paths_config", type=str, default="configs/paths.yaml")
    ap.add_argument("--dataset-config", dest="dataset_config", type=str, default=None)
    return ap.parse_args()


def main() -> int:
    args = parse_args()
    repo_root = Path(__file__).resolve().parents[1]
    paths_cfg = load_paths_config(args.paths_config, repo_root=repo_root)
    roots = resolve_dataset_roots(paths_cfg, dataset=args.dataset)
    dataset_cfg = load_dataset_config(paths_cfg, dataset_config_path=args.dataset_config, repo_root=repo_root)

    behavioral_cfg = dataset_cfg.get("behavioral", {})
    app_data_rel = behavioral_cfg.get("app_data_dir")
    seq_rel = behavioral_cfg.get("app_sequence_dir")
    corrected_rel = behavioral_cfg.get("corrected_app_excel_dir")
    interim_rel = behavioral_cfg.get("interim_preprocess_dir", "behavioral_preprocess")
    if not app_data_rel or not seq_rel or not corrected_rel:
        raise ValueError("Missing dataset.behavioral.{app_data_dir,app_sequence_dir,corrected_app_excel_dir}.")

    raw_app_dir = roots["raw_root"] / app_data_rel
    seq_root = roots["raw_root"] / seq_rel
    interim_dir = roots["interim_root"] / interim_rel
    out_dir = roots["processed_root"] / corrected_rel / "run_corrected_v2" / "sequence_library"
    out_dir.mkdir(parents=True, exist_ok=True)

    item_group2 = _read_subject_ids(interim_dir / "groups_by_item" / "group_002_sublist.txt")
    answer_group2 = _read_subject_ids(interim_dir / "groups_by_answer" / "group_002_sublist.txt")
    raw_index = _index_raw_workbooks(raw_app_dir)

    item_templates = _build_templates_from_subjects(
        subject_ids=item_group2,
        raw_index=raw_index,
        target_col=ITEM_COL,
    )
    answer_templates = _build_templates_from_subjects(
        subject_ids=answer_group2,
        raw_index=raw_index,
        target_col=ANSWER_COL,
    )

    visit3_groups: dict[str, Any] = {}
    for task in sorted(set(item_templates) | set(answer_templates)):
        it = item_templates.get(task)
        at = answer_templates.get(task)
        visit3_groups[task] = {
            "items_source": "item_group_002_template" if it else None,
            "answers_source": "answer_group_002_template" if at else None,
            "item_template_subject_id": (it.subject_id if it else None),
            "answer_template_subject_id": (at.subject_id if at else None),
            "item_template_file_name": (it.file_name if it else None),
            "answer_template_file_name": (at.file_name if at else None),
            "items_n_rows": (it.n_rows if it else None),
            "answers_n_rows": (at.n_rows if at else None),
            "items": (it.sequence if it else None),
            "answers": (at.sequence if at else None),
        }

    visit3_app_seq = _read_app_sequence_visit(seq_root, "visit3")
    compare_rows = _compare_to_app_sequence(visit3_groups=visit3_groups, visit3_app_seq=visit3_app_seq)

    out_items = {
        task: {
            "subject_id": t.subject_id,
            "file_name": t.file_name,
            "n_rows": t.n_rows,
            "items": t.sequence,
        }
        for task, t in sorted(item_templates.items())
    }
    out_answers = {
        task: {
            "subject_id": t.subject_id,
            "file_name": t.file_name,
            "n_rows": t.n_rows,
            "answers": t.sequence,
        }
        for task, t in sorted(answer_templates.items())
    }

    (out_dir / "visit3_item_group2_templates.json").write_text(
        json.dumps(out_items, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (out_dir / "visit3_answer_group2_templates.json").write_text(
        json.dumps(out_answers, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (out_dir / "visit3_from_groups.json").write_text(
        json.dumps(visit3_groups, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (out_dir / "visit3_from_app_sequence.json").write_text(
        json.dumps(visit3_app_seq, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (out_dir / "visit3_compare_to_app_sequence.json").write_text(
        json.dumps(compare_rows, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(f"item_group2_subjects={len(item_group2)}")
    print(f"answer_group2_subjects={len(answer_group2)}")
    print(f"item_templates_tasks={len(item_templates)}")
    print(f"answer_templates_tasks={len(answer_templates)}")
    print(f"app_sequence_visit3_tasks={len(visit3_app_seq)}")
    print(f"out_dir={out_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

