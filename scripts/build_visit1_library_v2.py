from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

from src.path_config import load_dataset_config, load_paths_config, resolve_dataset_roots


ITEM_COL = "正式阶段刺激图片/Item名"
ANSWER_COL = "正式阶段正确答案"


def _normalize_header(value: object) -> str:
    s = str(value).strip()
    return "".join(s.split())


def _find_col_idx(header_row: list[object], target: str) -> int | None:
    target_norm = _normalize_header(target)
    for idx, val in enumerate(header_row, start=1):
        if val is None:
            continue
        if _normalize_header(val) == target_norm:
            return idx
    return None


def _read_column(ws: openpyxl.worksheet.worksheet.Worksheet, col_idx: int, *, max_row: int | None = None) -> list[Any]:
    mr = max_row or ws.max_row
    vals: list[Any] = []
    for row in ws.iter_rows(min_row=2, max_row=mr, min_col=col_idx, max_col=col_idx, values_only=True):
        vals.append(row[0])
    return vals


def _canonical_task_name(raw: str) -> str:
    s = str(raw).strip()
    if not s:
        return s
    rules = {
        "LG": "EmotionSwitch",
        "STROOP": "ColorStroop",
        "SpatialNBack": "Spatial2Back",
        "Emotion2Backformal": "Emotion2Back",
        "Number2Backformal": "Number2Back",
        "Spatial1Backformal": "Spatial1Back",
        "EmotionStoop": "EmotionStroop",
    }
    if s in rules:
        return rules[s]
    if "formal" in s.lower():
        s = re.sub("formal", "", s, flags=re.IGNORECASE).strip()
    return s


def _extract_subject_id_from_filename(excel_path: Path) -> str:
    stem = excel_path.stem
    stem = stem.replace("GameData", "").rstrip("_- ")
    parts = [p for p in stem.split("_") if p]
    if len(parts) >= 4:
        return "_".join(parts[:4])
    return stem


def _resolve_app_paths(*, roots: dict[str, Path], dataset_cfg: dict[str, Any]) -> tuple[Path, Path, Path]:
    behavioral_cfg = dataset_cfg.get("behavioral", {})
    app_data_rel = behavioral_cfg.get("app_data_dir")
    seq_rel = behavioral_cfg.get("app_sequence_dir")
    corrected_rel = behavioral_cfg.get("corrected_app_excel_dir")
    interim_rel = behavioral_cfg.get("interim_preprocess_dir", "behavioral_preprocess")
    if not app_data_rel or not seq_rel or not corrected_rel:
        raise ValueError("Missing dataset.behavioral.{app_data_dir,app_sequence_dir,corrected_app_excel_dir}.")
    raw_app_dir = roots["raw_root"] / app_data_rel
    seq_root = roots["raw_root"] / seq_rel
    corrected_root = roots["processed_root"] / corrected_rel
    interim_dir = roots["interim_root"] / interim_rel
    return raw_app_dir, seq_root, corrected_root / "run_corrected_v2" / "sequence_library"


def _read_group1_subject_ids(interim_dir: Path) -> list[str]:
    groups_manifest = interim_dir / "groups_by_item" / "groups_manifest.json"
    if not groups_manifest.exists():
        raise FileNotFoundError(f"Missing groups manifest: {groups_manifest}")
    rows = json.loads(groups_manifest.read_text(encoding="utf-8"))
    group1 = next((r for r in rows if r.get("group_id") == "group_001"), None)
    if not group1:
        raise ValueError("Missing group_001 in groups_manifest.json")
    subjects = str(group1.get("subjects") or "").split(";")
    subjects = [s for s in (x.strip() for x in subjects) if s]
    if not subjects:
        raise ValueError("group_001 has empty subject list")
    return subjects


def _index_raw_workbooks(raw_app_dir: Path) -> dict[str, Path]:
    out: dict[str, Path] = {}
    for fp in raw_app_dir.glob("*.xlsx"):
        if fp.name.startswith("~$"):
            continue
        sid = _extract_subject_id_from_filename(fp)
        # If duplicates exist, keep the first seen (raw dir should already be de-duplicated).
        out.setdefault(sid, fp)
    return out


def _read_app_sequence_visit1(seq_root: Path) -> dict[str, dict[str, list[Any]]]:
    """
    Parse visit1 configs into per-task sequences.

    Output:
      task -> {"items": [...], "answers": [...], "source": "<path>"}
    """
    visit1_dir = seq_root / "visit1"
    out: dict[str, dict[str, list[Any]]] = {}
    for p in sorted(visit1_dir.glob("*")):
        if p.is_dir() or p.suffix.lower() not in {".json", ".txt"}:
            continue
        try:
            obj = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            try:
                obj = json.loads(p.read_text(encoding="gbk"))
            except Exception:
                continue
        if not isinstance(obj, dict):
            continue
        items_key = next((k for k in obj.keys() if isinstance(k, str) and k.lower().endswith("_items")), None)
        if not items_key:
            continue
        raw_items = obj.get(items_key)
        if not isinstance(raw_items, list):
            continue

        task = items_key
        task = re.sub(r"_items$", "", task, flags=re.IGNORECASE)
        task = re.sub(r"formal\d*$", "", task, flags=re.IGNORECASE)
        task = _canonical_task_name(task).strip("_- ")

        items: list[Any] = []
        answers: list[Any] = []
        for rec in raw_items:
            if not isinstance(rec, dict):
                continue

            # items: standard tasks (PicName-like keys)
            item_val = None
            for k in ("step2_PicName", "step1_PicName", "PicName", "picName", "picData", "PicData"):
                if k in rec and rec[k] is not None:
                    item_val = rec[k]
                    break

            # answers: prefer buttonName; fallback to needClickButton for CPT/GNG; fallback to nested answerPicName list.
            ans_val = None
            for k in ("buttonName", "ButtonName"):
                if k in rec and rec[k] is not None:
                    ans_val = rec[k]
                    break
            if ans_val is None and "needClickButton" in rec:
                ans_val = rec.get("needClickButton")

            # KT/ZYST-like configs: nested answers list
            if ans_val is None and isinstance(rec.get("answers"), list):
                for a in rec["answers"]:
                    if isinstance(a, dict) and a.get("answerPicName") is not None:
                        answers.append(a.get("answerPicName"))

            if item_val is not None:
                items.append(item_val)
            # If we already appended nested answers, do not append ans_val for that record.
            if ans_val is not None:
                answers.append(ans_val)

        out[task] = {"items": items, "answers": answers, "source": [p.as_posix()]}
    return out


@dataclass(frozen=True)
class Visit1Template:
    task: str
    subject_id: str
    file_name: str
    n_rows: int
    items: list[Any] | None
    answers: list[Any] | None


def _build_visit1_templates_from_group1(
    *,
    group1_subjects: list[str],
    raw_index: dict[str, Path],
) -> dict[str, Visit1Template]:
    """
    Infer per-task column templates from item group_001 subjects.

    We search group1 in order and take the first workbook that provides the task
    with the relevant columns. This is sufficient because group_001 is expected
    to be internally consistent on item sequences for overlapping tasks.
    """
    templates: dict[str, Visit1Template] = {}
    for sid in group1_subjects:
        fp = raw_index.get(sid)
        if fp is None:
            continue
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        try:
            for sheet_name in wb.sheetnames:
                task = _canonical_task_name(sheet_name)
                if task in templates:
                    continue
                ws = wb[sheet_name]
                try:
                    header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
                except StopIteration:
                    continue
                item_idx = _find_col_idx(header, ITEM_COL)
                ans_idx = _find_col_idx(header, ANSWER_COL)
                if item_idx is None and ans_idx is None:
                    continue

                max_row = ws.max_row
                if task.upper() == "SST" and max_row == 98:
                    max_row = 97  # ignore the known invalid last row

                items = _read_column(ws, item_idx, max_row=max_row) if item_idx is not None else None
                answers = _read_column(ws, ans_idx, max_row=max_row) if ans_idx is not None else None
                templates[task] = Visit1Template(
                    task=task,
                    subject_id=sid,
                    file_name=fp.name,
                    n_rows=max_row - 1,
                    items=items,
                    answers=answers,
                )
        finally:
            wb.close()

    return templates


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(
        description=(
            "Build a visit1 sequence library for v2: infer visit1 item templates from item group_001 "
            "and parse visit1 items/answers from app_sequence configs."
        )
    )
    ap.add_argument("--dataset", type=str, required=True)
    ap.add_argument("--config", dest="paths_config", type=str, default="configs/paths.yaml")
    ap.add_argument("--dataset-config", dest="dataset_config", type=str, default=None)
    return ap.parse_args()


def main() -> int:
    args = parse_args()
    repo_root = Path(__file__).resolve().parents[1]
    paths_cfg = load_paths_config(args.paths_config, repo_root=repo_root)
    roots = resolve_dataset_roots(paths_cfg, dataset=args.dataset)
    dataset_cfg = load_dataset_config(paths_cfg, dataset_config_path=args.dataset_config, repo_root=repo_root)

    behavioral_cfg = dataset_cfg.get("behavioral", {})
    interim_dir = roots["interim_root"] / behavioral_cfg.get("interim_preprocess_dir", "behavioral_preprocess")

    raw_app_dir, seq_root, out_dir = _resolve_app_paths(roots=roots, dataset_cfg=dataset_cfg)
    out_dir.mkdir(parents=True, exist_ok=True)

    group1_subjects = _read_group1_subject_ids(interim_dir)
    raw_index = _index_raw_workbooks(raw_app_dir)

    templates = _build_visit1_templates_from_group1(group1_subjects=group1_subjects, raw_index=raw_index)
    seqs = _read_app_sequence_visit1(seq_root)

    merged: dict[str, Any] = {}
    for task in sorted(set(templates) | set(seqs)):
        t = templates.get(task)
        s = seqs.get(task)
        items = t.items if t else None
        answers = None
        if s and s.get("answers"):
            answers = s.get("answers")
        elif t:
            answers = t.answers
        note = None
        if t and answers is not None and isinstance(answers, list) and t.n_rows != len(answers):
            note = f"length_mismatch: template_n_rows={t.n_rows} answers_len={len(answers)}"
        merged[task] = {
            "items_source": "item_group1_template" if t else None,
            "answers_source": "app_sequence_visit1" if (s and s.get('answers')) else ("item_group1_template" if t else None),
            "template_subject_id": t.subject_id if t else None,
            "template_file_name": t.file_name if t else None,
            "n_rows": t.n_rows if t else None,
            "items": items,
            "answers": answers,
            "note": note,
            "app_sequence_source": (s.get("source") if s else None),
        }

    # Persist
    out_templates = {
        task: {
            "subject_id": t.subject_id,
            "file_name": t.file_name,
            "n_rows": t.n_rows,
            "items": t.items,
            "answers": t.answers,
        }
        for task, t in sorted(templates.items())
    }
    (out_dir / "visit1_item_group1_templates.json").write_text(
        json.dumps(out_templates, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    (out_dir / "visit1_from_app_sequence.json").write_text(
        json.dumps(seqs, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    (out_dir / "visit1_merged.json").write_text(
        json.dumps(merged, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    print(f"raw_workbooks_indexed={len(raw_index)}")
    print(f"group1_subjects={len(group1_subjects)}")
    print(f"templates_tasks={len(templates)}")
    print(f"visit1_app_sequence_tasks={len(seqs)}")
    print(f"out_dir={out_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
