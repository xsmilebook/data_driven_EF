from __future__ import annotations

import argparse
import json
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

from src.path_config import load_dataset_config, load_paths_config, resolve_dataset_roots


ITEM_COL = "正式阶段刺激图片/Item名"


def _normalize_header(value: object) -> str:
    s = str(value).strip()
    return "".join(s.split())


def _find_item_col_idx(header_row: list[object]) -> int | None:
    target_norm = _normalize_header(ITEM_COL)
    for idx, val in enumerate(header_row, start=1):
        if val is None:
            continue
        hv = _normalize_header(val)
        if hv == target_norm:
            return idx
    for idx, val in enumerate(header_row, start=1):
        if val is None:
            continue
        hv = _normalize_header(val).lower()
        if ("刺激" in hv or "pic" in hv) and ("item" in hv):
            return idx
    return None


def _canonical_task_name(raw: str) -> str:
    s = str(raw).strip()
    if not s:
        return s
    s = s.strip("_- ")
    rules = {
        "EmotionStoop": "EmotionStroop",
        "LG": "EmotionSwitch",
        "STROOP": "ColorStroop",
        "SpatialNBack": "Spatial2Back",
        "Emotion2Backformal": "Emotion2Back",
        "Number2Backformal": "Number2Back",
        "Spatial1Backformal": "Spatial1Back",
    }
    if s in rules:
        return rules[s]
    if "formal" in s.lower():
        s = re.sub("formal", "", s, flags=re.IGNORECASE).strip()
    return s


def _extract_subject_id_from_filename(excel_path: Path) -> str:
    stem = excel_path.stem.replace("GameData", "").rstrip("_- ")
    parts = [p for p in stem.split("_") if p]
    if len(parts) >= 4:
        return "_".join(parts[:4])
    return stem


def _index_raw_workbooks(raw_app_dir: Path) -> dict[str, Path]:
    out: dict[str, Path] = {}
    for fp in raw_app_dir.glob("*.xlsx"):
        if fp.name.startswith("~$"):
            continue
        sid = _extract_subject_id_from_filename(fp)
        out.setdefault(sid, fp)
    return out


def _load_item_templates(path: Path) -> dict[str, list[Any]]:
    obj = json.loads(path.read_text(encoding="utf-8"))
    out: dict[str, list[Any]] = {}
    for task, rec in obj.items():
        if isinstance(rec, dict) and isinstance(rec.get("items"), list):
            out[task] = rec["items"]
    return out


def _replace_items_in_workbook(xlsx: Path, item_templates: dict[str, list[Any]]) -> dict[str, Any]:
    wb = openpyxl.load_workbook(xlsx)
    updated_tasks = 0
    missing_item_col: list[str] = []
    missing_template: list[str] = []
    truncated_tasks: list[str] = []
    try:
        for sheet_name in wb.sheetnames:
            task = _canonical_task_name(sheet_name)
            tpl = item_templates.get(task)
            if tpl is None:
                missing_template.append(task)
                continue

            ws = wb[sheet_name]
            header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            item_idx = _find_item_col_idx(header)
            if item_idx is None:
                missing_item_col.append(task)
                continue

            data_rows = ws.max_row - 1
            n_write = min(data_rows, len(tpl))
            for i in range(n_write):
                ws.cell(row=2 + i, column=item_idx).value = tpl[i]
            if data_rows > len(tpl):
                for i in range(len(tpl), data_rows):
                    ws.cell(row=2 + i, column=item_idx).value = None
            if data_rows < len(tpl):
                truncated_tasks.append(task)
            updated_tasks += 1
        wb.save(xlsx)
    finally:
        wb.close()

    return {
        "updated_tasks": updated_tasks,
        "missing_item_col_tasks": sorted(set(missing_item_col)),
        "missing_template_tasks": sorted(set(missing_template)),
        "truncated_tasks": sorted(set(truncated_tasks)),
    }


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(description="Apply visit item templates to one raw APP workbook by subject_id.")
    ap.add_argument("--dataset", required=True)
    ap.add_argument("--subject-id", required=True)
    ap.add_argument("--visit-template", default="visit3_from_groups.json")
    ap.add_argument("--config", dest="paths_config", default="configs/paths.yaml")
    ap.add_argument("--dataset-config", dest="dataset_config", default=None)
    return ap.parse_args()


def main() -> int:
    args = parse_args()
    repo_root = Path(__file__).resolve().parents[1]
    paths_cfg = load_paths_config(args.paths_config, repo_root=repo_root)
    roots = resolve_dataset_roots(paths_cfg, dataset=args.dataset)
    dataset_cfg = load_dataset_config(paths_cfg, dataset_config_path=args.dataset_config, repo_root=repo_root)

    behavioral_cfg = dataset_cfg.get("behavioral", {})
    app_data_rel = behavioral_cfg.get("app_data_dir")
    corrected_rel = behavioral_cfg.get("corrected_app_excel_dir")
    if not app_data_rel or not corrected_rel:
        raise ValueError("Missing dataset.behavioral.app_data_dir/corrected_app_excel_dir.")

    raw_app_dir = roots["raw_root"] / app_data_rel
    seq_path = roots["processed_root"] / corrected_rel / "run_corrected_v2" / "sequence_library" / args.visit_template
    if not seq_path.exists():
        raise FileNotFoundError(f"Template not found: {seq_path}")

    workbook_index = _index_raw_workbooks(raw_app_dir)
    xlsx = workbook_index.get(args.subject_id)
    if xlsx is None:
        raise FileNotFoundError(f"Workbook for {args.subject_id} not found in {raw_app_dir}")

    templates = _load_item_templates(seq_path)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = raw_app_dir / f"_backup_item_replace_{args.subject_id}_{ts}"
    backup_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(xlsx, backup_dir / xlsx.name)

    info = _replace_items_in_workbook(xlsx, templates)
    print(f"subject_id={args.subject_id}")
    print(f"workbook={xlsx}")
    print(f"template={seq_path}")
    print(f"backup_dir={backup_dir}")
    print(f"updated_tasks={info['updated_tasks']}")
    print(f"missing_item_col_tasks={','.join(info['missing_item_col_tasks'])}")
    print(f"missing_template_tasks={','.join(info['missing_template_tasks'])}")
    print(f"truncated_tasks={','.join(info['truncated_tasks'])}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
