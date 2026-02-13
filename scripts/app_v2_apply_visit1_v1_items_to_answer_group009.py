from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

from src.path_config import load_dataset_config, load_paths_config, resolve_dataset_roots


ITEM_COL = "正式阶段刺激图片/Item名"


def _normalize_header(value: object) -> str:
    s = str(value).strip()
    return "".join(s.split())


def _find_item_col_idx(header_row: list[object]) -> int | None:
    target_norm = _normalize_header(ITEM_COL)
    for idx, val in enumerate(header_row, start=1):
        if val is None:
            continue
        hv = _normalize_header(val)
        if hv == target_norm:
            return idx
    for idx, val in enumerate(header_row, start=1):
        if val is None:
            continue
        hv = _normalize_header(val).lower()
        if ("刺激" in hv or "pic" in hv) and ("item" in hv):
            return idx
    return None


def _canonical_task_name(raw: str) -> str:
    s = str(raw).strip().strip("_- ")
    rules = {
        "EmotionStoop": "EmotionStroop",
        "LG": "EmotionSwitch",
        "STROOP": "ColorStroop",
        "SpatialNBack": "Spatial2Back",
        "Emotion2Backformal": "Emotion2Back",
        "Number2Backformal": "Number2Back",
        "Spatial1Backformal": "Spatial1Back",
    }
    if s in rules:
        return rules[s]
    if "formal" in s.lower():
        s = re.sub("formal", "", s, flags=re.IGNORECASE).strip()
    return s


def _extract_subject_id_from_filename(excel_path: Path) -> str:
    stem = excel_path.stem.replace("GameData", "").rstrip("_- ")
    parts = [p for p in stem.split("_") if p]
    if len(parts) >= 4:
        return "_".join(parts[:4])
    return stem


def _index_raw_workbooks(raw_app_dir: Path) -> dict[str, Path]:
    out: dict[str, Path] = {}
    for fp in raw_app_dir.glob("*.xlsx"):
        if fp.name.startswith("~$"):
            continue
        sid = _extract_subject_id_from_filename(fp)
        out.setdefault(sid, fp)
    return out


def _read_subject_list(path: Path) -> list[str]:
    return [s.strip() for s in path.read_text(encoding="utf-8").splitlines() if s.strip()]


def _load_item_templates(path: Path) -> dict[str, list[Any]]:
    obj = json.loads(path.read_text(encoding="utf-8"))
    out: dict[str, list[Any]] = {}
    for task, rec in obj.items():
        if isinstance(rec, dict) and isinstance(rec.get("items"), list):
            out[task] = rec["items"]
    return out


def _replace_items_in_workbook(xlsx: Path, item_templates: dict[str, list[Any]]) -> dict[str, Any]:
    wb = openpyxl.load_workbook(xlsx)
    updated_tasks = 0
    missing_item_col: list[str] = []
    missing_template: list[str] = []
    truncated_tasks: list[str] = []
    try:
        for sheet_name in wb.sheetnames:
            task = _canonical_task_name(sheet_name)
            tpl = item_templates.get(task)
            if tpl is None:
                missing_template.append(task)
                continue

            ws = wb[sheet_name]
            header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            item_idx = _find_item_col_idx(header)
            if item_idx is None:
                missing_item_col.append(task)
                continue

            data_rows = ws.max_row - 1
            n_write = min(data_rows, len(tpl))
            for i in range(n_write):
                ws.cell(row=2 + i, column=item_idx).value = tpl[i]
            if data_rows > len(tpl):
                for i in range(len(tpl), data_rows):
                    ws.cell(row=2 + i, column=item_idx).value = None
            if data_rows < len(tpl):
                truncated_tasks.append(task)
            updated_tasks += 1
        wb.save(xlsx)
    finally:
        wb.close()

    return {
        "updated_tasks": updated_tasks,
        "missing_item_col_tasks": sorted(set(missing_item_col)),
        "missing_template_tasks": sorted(set(missing_template)),
        "truncated_tasks": sorted(set(truncated_tasks)),
    }


def _write_groups_manifest(groups_dir: Path) -> None:
    group_files = sorted(groups_dir.glob("group_*_sublist.txt"))
    rows: list[dict[str, Any]] = []
    for p in group_files:
        gid = p.stem.replace("_sublist", "")
        sids = sorted([s.strip() for s in p.read_text(encoding="utf-8").splitlines() if s.strip()])
        rows.append(
            {
                "group_id": gid,
                "n_subjects": len(sids),
                "subjects": ";".join(sids),
                "grouping": "item",
                "group_column": ITEM_COL,
                "compare_mode": "sequence",
                "n_sheets_constraints": 18,
                "example_subject": (sids[0] if sids else ""),
                "note": ("manual_group_assignment" if gid in ("group_005", "group_006") else None),
            }
        )

    (groups_dir / "groups_manifest.json").write_text(
        json.dumps(rows, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    header = [
        "group_id",
        "n_subjects",
        "subjects",
        "grouping",
        "group_column",
        "compare_mode",
        "n_sheets_constraints",
        "example_subject",
        "note",
    ]
    with (groups_dir / "groups_manifest.csv").open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=header)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(
        description=(
            "Apply visit1_v1_confirmed item templates to all answer_group_009 subjects, "
            "then move these subjects from item group_003 to group_006."
        )
    )
    ap.add_argument("--dataset", required=True)
    ap.add_argument("--config", dest="paths_config", default="configs/paths.yaml")
    ap.add_argument("--dataset-config", dest="dataset_config", default=None)
    return ap.parse_args()


def main() -> int:
    args = parse_args()
    repo_root = Path(__file__).resolve().parents[1]
    paths_cfg = load_paths_config(args.paths_config, repo_root=repo_root)
    roots = resolve_dataset_roots(paths_cfg, dataset=args.dataset)
    dataset_cfg = load_dataset_config(paths_cfg, dataset_config_path=args.dataset_config, repo_root=repo_root)

    behavioral_cfg = dataset_cfg.get("behavioral", {})
    app_data_rel = behavioral_cfg.get("app_data_dir")
    corrected_rel = behavioral_cfg.get("corrected_app_excel_dir")
    interim_rel = behavioral_cfg.get("interim_preprocess_dir", "behavioral_preprocess")
    if not app_data_rel or not corrected_rel:
        raise ValueError("Missing app_data_dir/corrected_app_excel_dir in config.")

    raw_app_dir = roots["raw_root"] / app_data_rel
    interim_dir = roots["interim_root"] / interim_rel
    conformed_dir = roots["processed_root"] / corrected_rel / "run_corrected_v2" / "conformed"

    group009 = _read_subject_list(interim_dir / "groups_by_answer" / "group_009_sublist.txt")
    item_group3_path = interim_dir / "groups_by_item" / "group_003_sublist.txt"
    item_group6_path = interim_dir / "groups_by_item" / "group_006_sublist.txt"
    item_group3 = _read_subject_list(item_group3_path)
    item_group6 = _read_subject_list(item_group6_path) if item_group6_path.exists() else []

    item_templates = _load_item_templates(conformed_dir / "visit1_v1_confirmed.json")
    raw_index = _index_raw_workbooks(raw_app_dir)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_dir = raw_app_dir / f"_backup_item_replace_answer_group009_visit1v1_{ts}"
    backup_dir.mkdir(parents=True, exist_ok=True)

    missing_workbooks: list[str] = []
    updated = 0
    for sid in group009:
        xlsx = raw_index.get(sid)
        if xlsx is None:
            missing_workbooks.append(sid)
            continue
        shutil.copy2(xlsx, backup_dir / xlsx.name)
        _replace_items_in_workbook(xlsx, item_templates)
        updated += 1

    move_set = set(group009)
    group3_new = sorted([s for s in item_group3 if s not in move_set])
    group6_new = sorted(set(item_group6).union(move_set))
    item_group3_path.write_text("\n".join(group3_new) + ("\n" if group3_new else ""), encoding="utf-8")
    item_group6_path.write_text("\n".join(group6_new) + ("\n" if group6_new else ""), encoding="utf-8")
    _write_groups_manifest(interim_dir / "groups_by_item")

    print(f"answer_group009_n={len(group009)}")
    print(f"updated_subjects_n={updated}")
    print(f"missing_workbooks_n={len(missing_workbooks)}")
    print(f"backup_dir={backup_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
