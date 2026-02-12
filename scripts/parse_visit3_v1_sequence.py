from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

import openpyxl

from src.path_config import load_dataset_config, load_paths_config, resolve_dataset_roots


ANSWER_COL = "正式阶段正确答案"
ITEM_COL = "正式阶段刺激图片/Item名"


def _normalize_cell(value: object) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    low = s.lower()
    if low in {"nan", "none"}:
        return None
    s = s.replace("\\", "/")
    if "/" in s:
        s = s.split("/")[-1]
    s = s.strip().lower()
    for ext in (".png", ".jpg", ".jpeg", ".bmp", ".gif", ".tif", ".tiff", ".webp"):
        if s.endswith(ext):
            s = s[: -len(ext)]
            break
    if s.isdigit():
        s2 = s.lstrip("0")
        s = s2 if s2 else "0"
    return s or None


def _canonical_task_name(raw: str) -> str:
    s = str(raw).strip()
    if not s:
        return s
    s = s.strip("_- ")
    rules = {
        "EmotionStoop": "EmotionStroop",
        "LG": "EmotionSwitch",
        "STROOP": "ColorStroop",
        "SpatialNBack": "Spatial2Back",
        "Emotion2Backformal": "Emotion2Back",
        "Number2Backformal": "Number2Back",
        "Spatial1Backformal": "Spatial1Back",
    }
    if s in rules:
        return rules[s]
    if "formal" in s.lower():
        s = re.sub("formal", "", s, flags=re.IGNORECASE).strip()
    return s


def _read_emotion2back_xlsx(path: Path) -> dict[str, list[Any]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        # Expected columns: picName, buttonName
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        try:
            pic_idx = header.index("picName") + 1
        except ValueError:
            pic_idx = 1
        try:
            btn_idx = header.index("buttonName") + 1
        except ValueError:
            btn_idx = 2

        items: list[Any] = []
        answers: list[Any] = []
        for r in range(2, ws.max_row + 1):
            items.append(ws.cell(row=r, column=pic_idx).value)
            answers.append(ws.cell(row=r, column=btn_idx).value)
        return {"items": items, "answers": answers}
    finally:
        wb.close()


def _read_spaticalnback_xlsx(path: Path) -> dict[str, dict[str, list[Any]]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out: dict[str, dict[str, list[Any]]] = {}
    try:
        # This file has Sheet1 with column B as item sequence, column C as answers.
        ws = wb.active
        items: list[Any] = []
        answers: list[Any] = []
        for r in range(2, ws.max_row + 1):
            items.append(ws.cell(row=r, column=2).value)
            answers.append(ws.cell(row=r, column=3).value)
        out["Spatial1Back"] = {"items": items, "answers": answers}
        out["Spatial2Back"] = {"items": items, "answers": answers}
        return out
    finally:
        wb.close()


def _read_group_subjects(path: Path) -> list[str]:
    return [s.strip() for s in path.read_text(encoding="utf-8").splitlines() if s.strip()]


def _resolve_xlsx(raw_dir: Path, subject_id: str) -> Path | None:
    matches = sorted(raw_dir.glob(f"{subject_id}_*_GameData.xlsx"))
    return matches[0] if matches else None


def _read_subject_answers(xlsx: Path, task: str) -> list[str]:
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    try:
        ws_name = None
        for name in wb.sheetnames:
            if _canonical_task_name(name) == task:
                ws_name = name
                break
        if ws_name is None:
            return []
        ws = wb[ws_name]
        header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        try:
            ans_idx = header.index(ANSWER_COL) + 1
        except ValueError:
            return []
        max_row = ws.max_row
        if task.upper() == "SST" and (max_row - 1) > 96:
            max_row = 97
        seq: list[str] = []
        for r in range(2, max_row + 1):
            v = _normalize_cell(ws.cell(row=r, column=ans_idx).value)
            if v is not None:
                seq.append(v)
        return seq
    finally:
        wb.close()


def _match_ratio(a: list[str], b: list[str]) -> float | None:
    n = min(len(a), len(b))
    if n == 0:
        return None
    same = sum(1 for i in range(n) if a[i] == b[i])
    return same / n


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(
        description="Parse visit3_v1 Excel files into JSON sequences and optionally compare to answer_group_004."
    )
    ap.add_argument("--dataset", type=str, required=True)
    ap.add_argument("--config", dest="paths_config", type=str, default="configs/paths.yaml")
    ap.add_argument("--dataset-config", dest="dataset_config", type=str, default=None)
    ap.add_argument("--subject-id", type=str, default=None, help="Optional subject id for comparison.")
    return ap.parse_args()


def main() -> int:
    args = parse_args()
    repo_root = Path(__file__).resolve().parents[1]
    paths_cfg = load_paths_config(args.paths_config, repo_root=repo_root)
    roots = resolve_dataset_roots(paths_cfg, dataset=args.dataset)
    dataset_cfg = load_dataset_config(paths_cfg, dataset_config_path=args.dataset_config, repo_root=repo_root)

    behavioral_cfg = dataset_cfg.get("behavioral", {})
    app_seq_rel = behavioral_cfg.get("app_sequence_dir")
    corrected_rel = behavioral_cfg.get("corrected_app_excel_dir")
    app_data_rel = behavioral_cfg.get("app_data_dir")
    interim_rel = behavioral_cfg.get("interim_preprocess_dir", "behavioral_preprocess")
    if not app_seq_rel or not corrected_rel or not app_data_rel:
        raise ValueError("Missing dataset.behavioral.app_sequence_dir/corrected_app_excel_dir/app_data_dir.")

    visit3_v1_dir = roots["raw_root"] / app_seq_rel / "visit3_v1"
    out_dir = roots["processed_root"] / corrected_rel / "run_corrected_v2" / "sequence_library"
    out_dir.mkdir(parents=True, exist_ok=True)

    emo_file = visit3_v1_dir / "emotion2back2.1.xlsx"
    spatial_file = visit3_v1_dir / "SpaticalNback_旧序列.xlsx"

    visit3_v1 = {
        "Emotion2Back": _read_emotion2back_xlsx(emo_file),
        **_read_spaticalnback_xlsx(spatial_file),
    }

    out_path = out_dir / "visit3_v1_parsed.json"
    out_path.write_text(json.dumps(visit3_v1, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    compare = None
    raw_dir = roots["raw_root"] / app_data_rel
    if args.subject_id:
        sid = args.subject_id
    else:
        group4 = roots["interim_root"] / interim_rel / "groups_by_answer" / "group_004_sublist.txt"
        sids = _read_group_subjects(group4)
        sid = sids[0] if sids else None
    if sid:
        xlsx = _resolve_xlsx(raw_dir, sid)
        if xlsx:
            compare = {}
            for task in ("Spatial1Back", "Spatial2Back", "Emotion2Back"):
                subj_ans = _read_subject_answers(xlsx, task)
                seq = visit3_v1.get(task, {}).get("answers") or []
                seq = [x for x in (_normalize_cell(v) for v in seq) if x is not None]
                compare[task] = {
                    "subject_id": sid,
                    "subject_len": len(subj_ans),
                    "visit3_v1_len": len(seq),
                    "match_ratio": _match_ratio(subj_ans, seq),
                }

    if compare is not None:
        cmp_path = out_dir / "visit3_v1_compare_to_group004.json"
        cmp_path.write_text(json.dumps(compare, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        print(f"compare_out={cmp_path}")

    print(f"out={out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

