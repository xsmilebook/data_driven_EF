from __future__ import annotations

import argparse
import json
import re
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

from src.path_config import load_dataset_config, load_paths_config, resolve_dataset_roots


ITEM_COL = "正式阶段刺激图片/Item名"


def _normalize_header(value: object) -> str:
    s = str(value).strip()
    return "".join(s.split())


def _find_col_idx(header_row: list[object], target: str) -> int | None:
    target_norm = _normalize_header(target)
    for idx, val in enumerate(header_row, start=1):
        if val is None:
            continue
        if _normalize_header(val) == target_norm:
            return idx
    return None


def _canonical_task_name(raw: str) -> str:
    s = str(raw).strip()
    if not s:
        return s
    s = s.strip("_- ")
    rules = {
        "EmotionStoop": "EmotionStroop",
        "LG": "EmotionSwitch",
        "STROOP": "ColorStroop",
        "SpatialNBack": "Spatial2Back",
        "Emotion2Backformal": "Emotion2Back",
        "Number2Backformal": "Number2Back",
        "Spatial1Backformal": "Spatial1Back",
    }
    if s in rules:
        return rules[s]
    if "formal" in s.lower():
        s = re.sub("formal", "", s, flags=re.IGNORECASE).strip()
    return s


def _extract_subject_id_from_filename(excel_path: Path) -> str:
    stem = excel_path.stem
    stem = stem.replace("GameData", "").rstrip("_- ")
    parts = [p for p in stem.split("_") if p]
    if len(parts) >= 4:
        return "_".join(parts[:4])
    return stem


def _index_raw_workbooks(raw_app_dir: Path) -> dict[str, Path]:
    out: dict[str, Path] = {}
    for fp in raw_app_dir.glob("*.xlsx"):
        if fp.name.startswith("~$"):
            continue
        sid = _extract_subject_id_from_filename(fp)
        # If duplicates exist, keep the first seen (raw dir should already be de-duplicated).
        out.setdefault(sid, fp)
    return out


def _read_subject_list(p: Path) -> list[str]:
    return [s.strip() for s in p.read_text(encoding="utf-8").splitlines() if s.strip()]


def _load_group1_item_templates(v2_seq_dir: Path) -> dict[str, list[Any] | None]:
    """
    Load per-task item templates inferred from item group_001.

    Returns:
      task -> items (list with possible None) or None if missing
    """
    obj = json.loads((v2_seq_dir / "visit1_item_group1_templates.json").read_text(encoding="utf-8"))
    out: dict[str, list[Any] | None] = {}
    for task, rec in obj.items():
        if not isinstance(rec, dict):
            continue
        out[task] = rec.get("items")
    return out


@dataclass(frozen=True)
class ReplaceResult:
    subject_id: str
    excel_file: str
    backup_file: str | None
    updated_tasks: int
    missing_item_col_tasks: list[str]
    missing_template_tasks: list[str]
    note: str | None = None


def _replace_items_in_workbook(
    *,
    excel_path: Path,
    templates: dict[str, list[Any] | None],
    dry_run: bool,
) -> ReplaceResult:
    sid = _extract_subject_id_from_filename(excel_path)
    missing_item_col: list[str] = []
    missing_template: list[str] = []
    updated = 0

    wb = openpyxl.load_workbook(excel_path)
    try:
        for sheet_name in wb.sheetnames:
            task = _canonical_task_name(sheet_name)
            tpl_items = templates.get(task)
            if tpl_items is None:
                missing_template.append(task)
                continue

            ws = wb[sheet_name]
            try:
                header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            except StopIteration:
                continue
            item_idx = _find_col_idx(header, ITEM_COL)
            if item_idx is None:
                missing_item_col.append(task)
                continue

            # If a known invalid last row exists for SST, mirror the template convention:
            # treat max_row=97 (96 trials) as the effective region for replacement.
            effective_max_row = ws.max_row
            if task.upper() == "SST" and effective_max_row == 98:
                effective_max_row = 97

            n_rows = max(0, effective_max_row - 1)
            if not dry_run:
                for i in range(n_rows):
                    row = 2 + i
                    val = tpl_items[i] if i < len(tpl_items) else None
                    ws.cell(row=row, column=item_idx).value = val
                # If we skipped SST invalid last row, clear its item as well for consistency.
                if task.upper() == "SST" and ws.max_row == 98:
                    ws.cell(row=98, column=item_idx).value = None
            updated += 1

        if not dry_run:
            wb.save(excel_path)
    finally:
        wb.close()

    return ReplaceResult(
        subject_id=sid,
        excel_file=excel_path.name,
        backup_file=None,
        updated_tasks=updated,
        missing_item_col_tasks=sorted(set(missing_item_col)),
        missing_template_tasks=sorted(set(missing_template)),
    )


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(
        description=(
            "Replace item sequences for subjects in answer group_002 that are not in item group_001, "
            "by overwriting the Excel item column with the item group_001 templates."
        )
    )
    ap.add_argument("--dataset", type=str, required=True)
    ap.add_argument("--config", dest="paths_config", type=str, default="configs/paths.yaml")
    ap.add_argument("--dataset-config", dest="dataset_config", type=str, default=None)
    ap.add_argument("--dry-run", action="store_true", help="Do not modify any Excel files.")
    ap.add_argument(
        "--no-backup",
        action="store_true",
        help="Do not create backup copies (NOT recommended).",
    )
    return ap.parse_args()


def main() -> int:
    args = parse_args()
    repo_root = Path(__file__).resolve().parents[1]
    paths_cfg = load_paths_config(args.paths_config, repo_root=repo_root)
    roots = resolve_dataset_roots(paths_cfg, dataset=args.dataset)
    dataset_cfg = load_dataset_config(paths_cfg, dataset_config_path=args.dataset_config, repo_root=repo_root)

    behavioral_cfg = dataset_cfg.get("behavioral", {})
    app_data_rel = behavioral_cfg.get("app_data_dir")
    corrected_rel = behavioral_cfg.get("corrected_app_excel_dir")
    interim_rel = behavioral_cfg.get("interim_preprocess_dir", "behavioral_preprocess")
    if not app_data_rel or not corrected_rel:
        raise ValueError("Missing dataset.behavioral.{app_data_dir,corrected_app_excel_dir}.")

    raw_app_dir = roots["raw_root"] / app_data_rel
    interim_dir = roots["interim_root"] / interim_rel
    v2_seq_dir = roots["processed_root"] / corrected_rel / "run_corrected_v2" / "sequence_library"

    answer_group2 = _read_subject_list(interim_dir / "groups_by_answer" / "group_002_sublist.txt")
    item_group1 = _read_subject_list(interim_dir / "groups_by_item" / "group_001_sublist.txt")
    targets = [sid for sid in answer_group2 if sid not in set(item_group1)]

    raw_index = _index_raw_workbooks(raw_app_dir)
    templates = _load_group1_item_templates(v2_seq_dir)

    backup_dir = None
    if not args.dry_run and not args.no_backup:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_dir = raw_app_dir / f"_backup_item_replace_{ts}"
        backup_dir.mkdir(parents=True, exist_ok=True)

    results: list[ReplaceResult] = []
    missing_workbooks: list[str] = []
    for sid in targets:
        fp = raw_index.get(sid)
        if fp is None:
            missing_workbooks.append(sid)
            continue
        if backup_dir is not None:
            dst = backup_dir / fp.name
            if not dst.exists():
                shutil.copy2(fp, dst)
            # ReplaceResult stores only name, not path.
            res = _replace_items_in_workbook(excel_path=fp, templates=templates, dry_run=args.dry_run)
            results.append(
                ReplaceResult(
                    subject_id=res.subject_id,
                    excel_file=res.excel_file,
                    backup_file=dst.name,
                    updated_tasks=res.updated_tasks,
                    missing_item_col_tasks=res.missing_item_col_tasks,
                    missing_template_tasks=res.missing_template_tasks,
                    note=res.note,
                )
            )
        else:
            res = _replace_items_in_workbook(excel_path=fp, templates=templates, dry_run=args.dry_run)
            results.append(res)

    # Write a machine-readable log under temp for quick inspection (not committed).
    temp_dir = repo_root / "temp"
    temp_dir.mkdir(parents=True, exist_ok=True)
    out_log = temp_dir / "app_v2_item_replace_answer_group2_log.json"
    out_log.write_text(
        json.dumps(
            {
                "dataset": args.dataset,
                "dry_run": bool(args.dry_run),
                "backup_dir": (backup_dir.as_posix() if backup_dir else None),
                "answer_group2_n": len(answer_group2),
                "item_group1_n": len(item_group1),
                "targets_n": len(targets),
                "missing_workbooks_n": len(missing_workbooks),
                "missing_workbooks": missing_workbooks,
                "results": [r.__dict__ for r in results],
            },
            ensure_ascii=False,
            indent=2,
        )
        + "\n",
        encoding="utf-8",
    )

    print(f"answer_group2_n={len(answer_group2)} item_group1_n={len(item_group1)} targets_n={len(targets)}")
    print(f"missing_workbooks_n={len(missing_workbooks)}")
    print(f"backup_dir={backup_dir}" if backup_dir else "backup_dir=None")
    print(f"log={out_log}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
