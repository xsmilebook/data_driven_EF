from __future__ import annotations

import argparse
import glob
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

from src.path_config import load_dataset_config, load_paths_config, resolve_dataset_roots


ANSWER_COL = "正式阶段正确答案"


def _normalize_header(value: object) -> str:
    s = str(value).strip()
    return "".join(s.split())


def _normalize_cell(value: object) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    low = s.lower()
    if low in {"nan", "none"}:
        return None
    s = s.replace("\\", "/")
    if "/" in s:
        s = s.split("/")[-1]
    s = s.strip().lower()
    for ext in (".png", ".jpg", ".jpeg", ".bmp", ".gif", ".tif", ".tiff", ".webp"):
        if s.endswith(ext):
            s = s[: -len(ext)]
            break
    if s.isdigit():
        s2 = s.lstrip("0")
        s = s2 if s2 else "0"
    return s or None


def _canonical_task_name(raw: str) -> str:
    s = str(raw).strip()
    if not s:
        return s
    s = s.strip("_- ")
    rules = {
        "EmotionStoop": "EmotionStroop",
        "LG": "EmotionSwitch",
        "STROOP": "ColorStroop",
        "SpatialNBack": "Spatial2Back",
        "Emotion2Backformal": "Emotion2Back",
        "Number2Backformal": "Number2Back",
        "Spatial1Backformal": "Spatial1Back",
    }
    if s in rules:
        return rules[s]
    if "formal" in s.lower():
        s = re.sub("formal", "", s, flags=re.IGNORECASE).strip()
    return s


def _extract_subject_id_from_filename(excel_path: Path) -> str:
    stem = excel_path.stem
    stem = stem.replace("GameData", "").rstrip("_- ")
    parts = [p for p in stem.split("_") if p]
    if len(parts) >= 4:
        return "_".join(parts[:4])
    return stem


def _find_col_idx(header_row: list[object], target: str) -> int | None:
    target_norm = _normalize_header(target)
    for idx, val in enumerate(header_row, start=1):
        if val is None:
            continue
        if _normalize_header(val) == target_norm:
            return idx
    return None


def _load_visit1_expected_answers(v2_seq_dir: Path) -> list[str | None]:
    merged = json.loads((v2_seq_dir / "visit1_merged.json").read_text(encoding="utf-8"))
    emo = merged.get("EmotionStroop")
    if not isinstance(emo, dict) or "answers" not in emo:
        raise ValueError("Missing EmotionStroop answers in visit1_merged.json")
    return emo.get("answers") or []


@dataclass(frozen=True)
class FixResult:
    subject_id: str
    excel_path: Path
    sst_truncated: bool
    sst_old_rows: int | None
    sst_new_rows: int | None
    emostroop_filled_trial13: bool
    match_key_tasks_to_group1_ref: bool | None
    match_note: str | None = None


def _truncate_sst_if_needed(wb: openpyxl.Workbook) -> tuple[bool, int | None, int | None]:
    # Find SST sheet by canonical name.
    sst_sheet = None
    for name in wb.sheetnames:
        if _canonical_task_name(name).upper() == "SST":
            sst_sheet = name
            break
    if sst_sheet is None:
        return False, None, None

    ws = wb[sst_sheet]
    data_rows = ws.max_row - 1
    if data_rows <= 96:
        return False, data_rows, data_rows

    # Keep header + first 96 trials (rows 2..97). Delete rows 98..end.
    delete_from = 98
    n_delete = ws.max_row - (delete_from - 1)
    if n_delete > 0:
        ws.delete_rows(delete_from, amount=n_delete)
    return True, data_rows, ws.max_row - 1


def _fill_emostroop_trial13_if_needed(
    wb: openpyxl.Workbook,
    *,
    expected_answers: list[str | None],
) -> bool:
    # Find EmotionStroop sheet by canonical name.
    sheet = None
    for name in wb.sheetnames:
        if _canonical_task_name(name) == "EmotionStroop":
            sheet = name
            break
    if sheet is None:
        return False

    ws = wb[sheet]
    try:
        header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    except StopIteration:
        return False
    ans_idx = _find_col_idx(header, ANSWER_COL)
    if ans_idx is None:
        return False

    # Trial index is 1-based; trial 13 corresponds to row 14 in the sheet.
    trial13_row = 14
    if ws.max_row < trial13_row:
        return False

    cur = ws.cell(row=trial13_row, column=ans_idx).value
    if _normalize_cell(cur) is not None:
        return False

    # Compare all other trials (excluding trial 13) to expected.
    # Use only up to the expected length.
    n = min(len(expected_answers), ws.max_row - 1)
    if n < 13:
        return False

    for trial in range(1, n + 1):
        if trial == 13:
            continue
        row = 1 + trial
        obs = _normalize_cell(ws.cell(row=row, column=ans_idx).value)
        exp = _normalize_cell(expected_answers[trial - 1])
        if obs != exp:
            return False

    ws.cell(row=trial13_row, column=ans_idx).value = "AN"
    return True


def _read_answer_sequences(
    wb: openpyxl.Workbook,
    *,
    expected_answers_emostroop: list[str | None],
) -> dict[str, list[str | None]]:
    """
    Read normalized answer sequences for grouping comparisons.
    SST is treated as first 96 trials only (after truncation step, if applied).
    """
    out: dict[str, list[str | None]] = {}
    for name in wb.sheetnames:
        task = _canonical_task_name(name)
        ws = wb[name]
        try:
            header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
        except StopIteration:
            continue
        ans_idx = _find_col_idx(header, ANSWER_COL)
        if ans_idx is None:
            continue

        max_row = ws.max_row
        if task.upper() == "SST" and (max_row - 1) > 96:
            max_row = 97

        seq: list[str | None] = []
        for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=ans_idx, max_col=ans_idx, values_only=True):
            seq.append(_normalize_cell(row[0]))
        out[task] = seq
    return out


def _write_groups_manifest(groups_dir: Path) -> None:
    group_files = sorted(groups_dir.glob("group_*_sublist.txt"))
    rows: list[dict[str, object]] = []
    for p in group_files:
        gid = p.stem.replace("_sublist", "")
        sids = [s.strip() for s in p.read_text(encoding="utf-8").splitlines() if s.strip()]
        rows.append(
            {
                "group_id": gid,
                "n_subjects": len(sids),
                "subjects": ";".join(sorted(sids)),
                "grouping": "answer",
                "group_column": "正式阶段正确答案",
                "compare_mode": "sequence",
                "example_subject": (sorted(sids)[0] if sids else ""),
                "note": None,
            }
        )
    # Write JSON/CSV for quick browsing (mirrors temp/group_app_stimulus_groups.py output schema loosely).
    (groups_dir / "groups_manifest.json").write_text(json.dumps(rows, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    # Minimal CSV without pandas dependency.
    header = [
        "group_id",
        "n_subjects",
        "subjects",
        "grouping",
        "group_column",
        "compare_mode",
        "example_subject",
        "note",
    ]
    lines = [",".join(header)]
    for r in rows:
        row = []
        for h in header:
            v = r.get(h, "")
            s = "" if v is None else str(v)
            # CSV escaping
            if any(ch in s for ch in [",", "\"", "\n"]):
                s = "\"" + s.replace("\"", "\"\"") + "\""
            row.append(s)
        lines.append(",".join(row))
    (groups_dir / "groups_manifest.csv").write_text("\n".join(lines) + "\n", encoding="utf-8")


def _load_group_subjects(groups_dir: Path, gid: str) -> list[str]:
    p = groups_dir / f"{gid}_sublist.txt"
    if not p.exists():
        raise FileNotFoundError(f"Missing group file: {p}")
    return [s.strip() for s in p.read_text(encoding="utf-8").splitlines() if s.strip()]


def _write_group_subjects(groups_dir: Path, gid: str, subjects: list[str]) -> None:
    p = groups_dir / f"{gid}_sublist.txt"
    p.write_text("\n".join(subjects) + ("\n" if subjects else ""), encoding="utf-8")


def _resolve_raw_workbook(raw_dir: Path, subject_id: str) -> Path | None:
    matches = glob.glob(str(raw_dir / f"{subject_id}_*_GameData.xlsx"))
    if not matches:
        return None
    return Path(matches[0])


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(
        description=(
            "Apply v2 fixes that should merge answer groups (SST truncation + EmotionStroop trial13 fill), "
            "then move subjects into answer group_001 if they match it."
        )
    )
    ap.add_argument("--dataset", type=str, required=True)
    ap.add_argument("--config", dest="paths_config", type=str, default="configs/paths.yaml")
    ap.add_argument("--dataset-config", dest="dataset_config", type=str, default=None)
    ap.add_argument("--dry-run", action="store_true", help="Do not write changes; only report what would change.")
    return ap.parse_args()


def main() -> int:
    args = parse_args()
    repo_root = Path(__file__).resolve().parents[1]
    paths_cfg = load_paths_config(args.paths_config, repo_root=repo_root)
    roots = resolve_dataset_roots(paths_cfg, dataset=args.dataset)
    ds_cfg = load_dataset_config(paths_cfg, dataset_config_path=args.dataset_config, repo_root=repo_root)

    behavioral_cfg = ds_cfg.get("behavioral", {})
    raw_app_rel = behavioral_cfg.get("app_data_dir")
    interim_rel = behavioral_cfg.get("interim_preprocess_dir", "behavioral_preprocess")
    corrected_rel = behavioral_cfg.get("corrected_app_excel_dir")
    if not raw_app_rel or not corrected_rel:
        raise ValueError("Missing dataset.behavioral.app_data_dir or corrected_app_excel_dir in config.")

    raw_dir = roots["raw_root"] / raw_app_rel
    groups_dir = roots["interim_root"] / interim_rel / "groups_by_answer"
    v2_seq_dir = roots["processed_root"] / corrected_rel / "run_corrected_v2" / "sequence_library"

    expected_emostroop = _load_visit1_expected_answers(v2_seq_dir)

    # Target groups to merge into group_001.
    g1 = "group_001"
    other_groups = ["group_003", "group_005", "group_006"]

    group1_subjects = _load_group_subjects(groups_dir, g1)
    target_subjects: list[tuple[str, str]] = []
    for gid in other_groups:
        for sid in _load_group_subjects(groups_dir, gid):
            target_subjects.append((sid, gid))

    if not group1_subjects:
        raise ValueError("group_001 is empty; cannot use it as merge target.")

    # Use the first subject in group_001 as the reference for overlap comparison.
    ref_sid = group1_subjects[0]
    ref_fp = _resolve_raw_workbook(raw_dir, ref_sid)
    if ref_fp is None:
        raise FileNotFoundError(f"Missing reference workbook for {ref_sid}")

    ref_wb = openpyxl.load_workbook(ref_fp, read_only=True, data_only=True)
    try:
        ref_seq = _read_answer_sequences(ref_wb, expected_answers_emostroop=expected_emostroop)
    finally:
        ref_wb.close()

    fixes: list[FixResult] = []
    moved: list[str] = []

    # Apply fixes + decide moves.
    for sid, from_gid in target_subjects:
        fp = _resolve_raw_workbook(raw_dir, sid)
        if fp is None:
            not_moved.append((sid, from_gid, "missing_workbook"))
            continue

        wb = openpyxl.load_workbook(fp)
        try:
            sst_trunc, old_rows, new_rows = _truncate_sst_if_needed(wb)
            emo_fill = _fill_emostroop_trial13_if_needed(wb, expected_answers=expected_emostroop)

            if (sst_trunc or emo_fill) and not args.dry_run:
                wb.save(fp)
            # Compare using the in-memory post-fix workbook (so --dry-run reflects intended effects).
            seq = _read_answer_sequences(wb, expected_answers_emostroop=expected_emostroop)

        finally:
            wb.close()

        # Audit signal: do the fixed workbook match group_001 reference on the key split tasks?
        key_tasks = {"SST", "EmotionStroop"}
        overlap = sorted((set(seq) & set(ref_seq) & key_tasks))
        ok: bool | None
        note: str | None = None
        if not overlap:
            ok = None
            note = "no_overlap_on_key_tasks(SST/EmotionStroop)"
        else:
            ok = True
            for task in overlap:
                if seq.get(task) != ref_seq.get(task):
                    ok = False
                    note = f"mismatch_task={task}"
                    break

        fixes.append(
            FixResult(
                subject_id=sid,
                excel_path=fp,
                sst_truncated=sst_trunc,
                sst_old_rows=old_rows,
                sst_new_rows=new_rows,
                emostroop_filled_trial13=emo_fill,
                match_key_tasks_to_group1_ref=ok,
                match_note=note,
            )
        )

        # User confirmed group_001/group_003/group_005/group_006 should be treated as the same answer group after fixes.
        # Merge all candidates into group_001; keep the key-task check only as an audit signal.
        if sid not in group1_subjects:
            group1_subjects.append(sid)
        moved.append(sid)

    # Update group lists
    group1_subjects = sorted(set(group1_subjects))
    if not args.dry_run:
        _write_group_subjects(groups_dir, g1, group1_subjects)

    for gid in other_groups:
        cur = _load_group_subjects(groups_dir, gid)
        cur2 = [sid for sid in cur if sid not in set(moved)]
        if not args.dry_run:
            _write_group_subjects(groups_dir, gid, cur2)
    if not args.dry_run:
        _write_groups_manifest(groups_dir)

    # Write a small report (Chinese) under docs/reports for audit.
    docs_reports_root = Path(paths_cfg.get("docs_reports_root", "docs/reports"))
    if not docs_reports_root.is_absolute():
        docs_reports_root = repo_root / docs_reports_root
    out_report = docs_reports_root / "app_v2_answer_group_merge.md"

    lines: list[str] = []
    lines.append("# APP v2: answer group merge fixes (SST/EmotionStroop)")
    lines.append("")
    lines.append("本次操作目的：将 `group_003/group_005/group_006` 中可通过规则修正的被试，合并回 `group_001`。")
    lines.append("")
    lines.append("## 规则")
    lines.append("")
    lines.append("- SST：若试次数 > 96，则截断为前 96 个 trial（删除第 96 个 trial 之后的行）。")
    lines.append(
        "- EmotionStroop：若第 13 个 trial 的 `正式阶段正确答案` 为空，且除第 13 个 trial 外其余答案与 "
        "`run_corrected_v2/sequence_library/visit1_merged.json` 一致，则用 `AN` 填充第 13 个 trial。"
    )
    lines.append("")
    lines.append("## 执行摘要")
    lines.append("")
    lines.append(f"- dry_run: {bool(args.dry_run)}")
    lines.append(f"- reference subject (group_001[0]): {ref_sid}")
    lines.append(f"- candidates: {len(target_subjects)}")
    lines.append(f"- moved_to_group_001: {len(moved)}")
    lines.append("")
    lines.append("备注：本次按人工确认将 group_003/group_005/group_006 全部合并进 group_001；")
    lines.append("下表中的 `match_key_tasks_to_group1_ref` 仅作为审计信号（SST/EmotionStroop 是否与 group_001 参考被试一致）。")
    lines.append("")
    lines.append("## 修正明细")
    lines.append("")
    lines.append("| subject_id | from_group | sst_truncated | sst_old_rows | sst_new_rows | emostroop_fill_trial13 | match_key_tasks_to_group1_ref | match_note |")
    lines.append("| --- | --- | --- | --- | --- | --- | --- | --- |")
    from_map = {sid: gid for sid, gid in target_subjects}
    for r in sorted(fixes, key=lambda x: x.subject_id):
        ok_s = "" if r.match_key_tasks_to_group1_ref is None else ("YES" if r.match_key_tasks_to_group1_ref else "NO")
        lines.append(
            f"| {r.subject_id} | {from_map.get(r.subject_id,'')} | {int(r.sst_truncated)} | "
            f"{'' if r.sst_old_rows is None else r.sst_old_rows} | {'' if r.sst_new_rows is None else r.sst_new_rows} | "
            f"{int(r.emostroop_filled_trial13)} | {ok_s} | {r.match_note or ''} |"
        )
    lines.append("")

    if not args.dry_run:
        out_report.write_text("\n".join(lines) + "\n", encoding="utf-8")
        print(f"wrote_report={out_report}")

    print(f"moved={len(moved)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
