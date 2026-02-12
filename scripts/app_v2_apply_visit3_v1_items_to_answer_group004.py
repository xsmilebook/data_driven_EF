from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

from src.path_config import load_dataset_config, load_paths_config, resolve_dataset_roots


ITEM_COL = "正式阶段刺激图片/Item名"


def _normalize_header(value: object) -> str:
    s = str(value).strip()
    return "".join(s.split())


def _find_item_col_idx(header_row: list[object]) -> int | None:
    target_norm = _normalize_header(ITEM_COL)
    for idx, val in enumerate(header_row, start=1):
        if val is None:
            continue
        hv = _normalize_header(val)
        if hv == target_norm:
            return idx
    # Fallback for unicode-variant headers.
    for idx, val in enumerate(header_row, start=1):
        if val is None:
            continue
        hv = _normalize_header(val).lower()
        if ("刺激" in hv or "pic" in hv) and ("item" in hv):
            return idx
    return None


def _canonical_task_name(raw: str) -> str:
    s = str(raw).strip()
    if not s:
        return s
    s = s.strip("_- ")
    rules = {
        "EmotionStoop": "EmotionStroop",
        "LG": "EmotionSwitch",
        "STROOP": "ColorStroop",
        "SpatialNBack": "Spatial2Back",
        "Emotion2Backformal": "Emotion2Back",
        "Number2Backformal": "Number2Back",
        "Spatial1Backformal": "Spatial1Back",
    }
    if s in rules:
        return rules[s]
    if "formal" in s.lower():
        s = re.sub("formal", "", s, flags=re.IGNORECASE).strip()
    return s


def _extract_subject_id_from_filename(excel_path: Path) -> str:
    stem = excel_path.stem
    stem = stem.replace("GameData", "").rstrip("_- ")
    parts = [p for p in stem.split("_") if p]
    if len(parts) >= 4:
        return "_".join(parts[:4])
    return stem


def _index_raw_workbooks(raw_app_dir: Path) -> dict[str, Path]:
    out: dict[str, Path] = {}
    for fp in raw_app_dir.glob("*.xlsx"):
        if fp.name.startswith("~$"):
            continue
        sid = _extract_subject_id_from_filename(fp)
        out.setdefault(sid, fp)
    return out


def _read_subject_list(path: Path) -> list[str]:
    return [s.strip() for s in path.read_text(encoding="utf-8").splitlines() if s.strip()]


def _load_visit3_v1_item_templates(path: Path) -> dict[str, list[Any]]:
    obj = json.loads(path.read_text(encoding="utf-8"))
    out: dict[str, list[Any]] = {}
    for task, rec in obj.items():
        if not isinstance(rec, dict):
            continue
        items = rec.get("items")
        if isinstance(items, list):
            out[task] = items
    return out


def _replace_items_in_workbook(
    *,
    xlsx: Path,
    item_templates: dict[str, list[Any]],
    dry_run: bool,
) -> dict[str, Any]:
    wb = openpyxl.load_workbook(xlsx)
    updated_tasks = 0
    missing_item_col: list[str] = []
    missing_template: list[str] = []
    truncated_tasks: list[str] = []
    try:
        for sheet_name in wb.sheetnames:
            task = _canonical_task_name(sheet_name)
            tpl = item_templates.get(task)
            if tpl is None:
                missing_template.append(task)
                continue

            ws = wb[sheet_name]
            header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            item_idx = _find_item_col_idx(header)
            if item_idx is None:
                missing_item_col.append(task)
                continue

            data_rows = ws.max_row - 1
            n_write = min(data_rows, len(tpl))
            if data_rows > len(tpl):
                # Clear overflow to avoid stale item residues.
                for i in range(len(tpl), data_rows):
                    ws.cell(row=2 + i, column=item_idx).value = None
            if data_rows < len(tpl):
                truncated_tasks.append(task)

            if not dry_run:
                for i in range(n_write):
                    ws.cell(row=2 + i, column=item_idx).value = tpl[i]
            updated_tasks += 1

        if not dry_run:
            wb.save(xlsx)
    finally:
        wb.close()

    return {
        "updated_tasks": updated_tasks,
        "missing_item_col_tasks": sorted(set(missing_item_col)),
        "missing_template_tasks": sorted(set(missing_template)),
        "truncated_tasks": sorted(set(truncated_tasks)),
    }


def _write_groups_manifest(groups_dir: Path) -> None:
    group_files = sorted(groups_dir.glob("group_*_sublist.txt"))
    rows: list[dict[str, Any]] = []
    for p in group_files:
        gid = p.stem.replace("_sublist", "")
        sids = sorted([s.strip() for s in p.read_text(encoding="utf-8").splitlines() if s.strip()])
        n_sheets_constraints = 18
        if gid == "group_004":
            n_sheets_constraints = 3
        rows.append(
            {
                "group_id": gid,
                "n_subjects": len(sids),
                "subjects": ";".join(sids),
                "grouping": "item",
                "group_column": ITEM_COL,
                "compare_mode": "sequence",
                "n_sheets_constraints": n_sheets_constraints,
                "example_subject": (sids[0] if sids else ""),
                "note": ("manual_group_assignment" if gid == "group_005" else None),
            }
        )

    (groups_dir / "groups_manifest.json").write_text(
        json.dumps(rows, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )

    header = [
        "group_id",
        "n_subjects",
        "subjects",
        "grouping",
        "group_column",
        "compare_mode",
        "n_sheets_constraints",
        "example_subject",
        "note",
    ]
    with (groups_dir / "groups_manifest.csv").open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=header)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def parse_args() -> argparse.Namespace:
    ap = argparse.ArgumentParser(
        description=(
            "Apply visit3_v1 item templates to all subjects in answer_group_004 (with backup), "
            "then move these subjects from item group_002 to item group_005."
        )
    )
    ap.add_argument("--dataset", type=str, required=True)
    ap.add_argument("--config", dest="paths_config", type=str, default="configs/paths.yaml")
    ap.add_argument("--dataset-config", dest="dataset_config", type=str, default=None)
    ap.add_argument("--dry-run", action="store_true")
    return ap.parse_args()


def main() -> int:
    args = parse_args()
    repo_root = Path(__file__).resolve().parents[1]
    paths_cfg = load_paths_config(args.paths_config, repo_root=repo_root)
    roots = resolve_dataset_roots(paths_cfg, dataset=args.dataset)
    dataset_cfg = load_dataset_config(paths_cfg, dataset_config_path=args.dataset_config, repo_root=repo_root)

    behavioral_cfg = dataset_cfg.get("behavioral", {})
    app_data_rel = behavioral_cfg.get("app_data_dir")
    corrected_rel = behavioral_cfg.get("corrected_app_excel_dir")
    interim_rel = behavioral_cfg.get("interim_preprocess_dir", "behavioral_preprocess")
    if not app_data_rel or not corrected_rel:
        raise ValueError("Missing dataset.behavioral.app_data_dir/corrected_app_excel_dir.")

    raw_app_dir = roots["raw_root"] / app_data_rel
    interim_dir = roots["interim_root"] / interim_rel
    seq_dir = roots["processed_root"] / corrected_rel / "run_corrected_v2" / "sequence_library"
    visit3_v1_path = seq_dir / "visit3_v1_from_groups.json"

    answer_group4 = _read_subject_list(interim_dir / "groups_by_answer" / "group_004_sublist.txt")
    item_group2_path = interim_dir / "groups_by_item" / "group_002_sublist.txt"
    item_group5_path = interim_dir / "groups_by_item" / "group_005_sublist.txt"

    item_group2 = _read_subject_list(item_group2_path)
    item_group5 = _read_subject_list(item_group5_path) if item_group5_path.exists() else []

    item_templates = _load_visit3_v1_item_templates(visit3_v1_path)
    raw_index = _index_raw_workbooks(raw_app_dir)

    backup_dir = None
    if not args.dry_run:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_dir = raw_app_dir / f"_backup_item_replace_answer_group004_visit3v1_{ts}"
        backup_dir.mkdir(parents=True, exist_ok=True)

    missing_workbooks: list[str] = []
    update_rows: list[dict[str, Any]] = []
    for sid in answer_group4:
        xlsx = raw_index.get(sid)
        if xlsx is None:
            missing_workbooks.append(sid)
            continue
        if backup_dir is not None:
            shutil.copy2(xlsx, backup_dir / xlsx.name)
        info = _replace_items_in_workbook(
            xlsx=xlsx,
            item_templates=item_templates,
            dry_run=args.dry_run,
        )
        info["subject_id"] = sid
        info["file_name"] = xlsx.name
        update_rows.append(info)

    # Update manual group assignment: group_002 -> group_005 for answer_group_004 subjects.
    if not args.dry_run:
        move_set = set(answer_group4)
        group2_new = sorted([s for s in item_group2 if s not in move_set])
        group5_new = sorted(set(item_group5).union(move_set))
        item_group2_path.write_text("\n".join(group2_new) + ("\n" if group2_new else ""), encoding="utf-8")
        item_group5_path.write_text("\n".join(group5_new) + ("\n" if group5_new else ""), encoding="utf-8")
        _write_groups_manifest(interim_dir / "groups_by_item")

    log = {
        "dry_run": bool(args.dry_run),
        "answer_group4_n": len(answer_group4),
        "item_group2_before_n": len(item_group2),
        "item_group5_before_n": len(item_group5),
        "missing_workbooks_n": len(missing_workbooks),
        "missing_workbooks": missing_workbooks,
        "backup_dir": (backup_dir.as_posix() if backup_dir else None),
        "updated_subjects_n": len(update_rows),
        "rows": update_rows,
    }

    temp_log = repo_root / "temp" / "app_v2_apply_visit3_v1_group004_log.json"
    temp_log.write_text(json.dumps(log, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    print(f"answer_group4_n={len(answer_group4)}")
    print(f"updated_subjects_n={len(update_rows)}")
    print(f"missing_workbooks_n={len(missing_workbooks)}")
    print(f"backup_dir={backup_dir}")
    print(f"log={temp_log}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

