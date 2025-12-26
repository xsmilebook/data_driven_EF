#!/usr/bin/env python3
"""
Script to rename 'EmotionStoop' sheet to 'EmotionStroop' in Excel files.
This script processes all .xlsx files in the specified directory.
"""

import os
import pandas as pd
import openpyxl
from pathlib import Path
import glob


def get_task_renaming_rules():
    """
    定义任务名称重命名规则
    
    Returns:
        dict: 原始名称 -> 目标名称的映射
    """
    return {
        'LG': 'EmotionSwitch',
        'STROOP': 'ColorStroop', 
        'SpatialNBack': 'Spatial2Back',
        'Emotion2Backformal': 'Emotion2Back',
        'Number2Backformal': 'Number2Back',
        'Spatial1Backformal': 'Spatial1Back',
        'EmotionStoop': 'EmotionStroop'  # 保留原有的修正
    }

def rename_sheets_in_excel(file_path):
    """
    根据重命名规则批量重命名Excel文件中的sheet
    
    Args:
        file_path (str): Path to the Excel file
    
    Returns:
        list: 被重命名的sheet列表
    """
    renaming_rules = get_task_renaming_rules()
    renamed_sheets = []
    
    try:
        # Load the workbook
        workbook = openpyxl.load_workbook(file_path)
        
        # 检查并应用重命名规则
        for old_name, new_name in renaming_rules.items():
            if old_name in workbook.sheetnames:
                # Get the sheet
                sheet = workbook[old_name]
                # Rename the sheet
                sheet.title = new_name
                renamed_sheets.append(f"'{old_name}' -> '{new_name}'")
                print(f"Renamed '{old_name}' to '{new_name}' in: {os.path.basename(file_path)}")
        
        # 处理其他带formal后缀的任务
        for sheet_name in workbook.sheetnames:
            if 'formal' in sheet_name.lower() and sheet_name not in renaming_rules:
                new_name = sheet_name.replace('formal', '')
                if new_name not in workbook.sheetnames:  # 确保新名称不存在
                    sheet = workbook[sheet_name]
                    sheet.title = new_name
                    renamed_sheets.append(f"'{sheet_name}' -> '{new_name}'")
                    print(f"Renamed '{sheet_name}' to '{new_name}' in: {os.path.basename(file_path)}")
        
        # Save the workbook if any changes were made
        if renamed_sheets:
            workbook.save(file_path)
            workbook.close()
            return renamed_sheets
        else:
            workbook.close()
            return []
            
    except Exception as e:
        print(f"Error processing {file_path}: {str(e)}")
        return []


def process_directory(directory_path):
    """
    Process all Excel files in the given directory.
    
    Args:
        directory_path (str): Path to the directory containing Excel files
    """
    # Get all .xlsx files in the directory
    excel_files = glob.glob(os.path.join(directory_path, "*.xlsx"))
    
    if not excel_files:
        print(f"No Excel files found in {directory_path}")
        return
    
    print(f"Found {len(excel_files)} Excel files to process")
    
    total_renamed_files = 0
    total_renamed_sheets = 0
    processed_count = 0
    
    for file_path in excel_files:
        processed_count += 1
        print(f"\nProcessing file {processed_count}/{len(excel_files)}: {os.path.basename(file_path)}")
        
        renamed_sheets = rename_sheets_in_excel(file_path)
        if renamed_sheets:
            total_renamed_files += 1
            total_renamed_sheets += len(renamed_sheets)
            print(f"  Renamed {len(renamed_sheets)} sheets: {', '.join(renamed_sheets)}")
    
    print(f"\nProcessing complete!")
    print(f"Total files processed: {processed_count}")
    print(f"Files with renamed sheets: {total_renamed_files}")
    print(f"Total sheets renamed: {total_renamed_sheets}")


def main():
    """Main function to run the script."""
    # Define the directory path
    directory_path = r"d:\code\data_driven_EF\data\EFNY\behavior_data\cibr_app_data"
    
    # Check if directory exists
    if not os.path.exists(directory_path):
        print(f"Error: Directory does not exist: {directory_path}")
        return
    
    if not os.path.isdir(directory_path):
        print(f"Error: Path is not a directory: {directory_path}")
        return
    
    print(f"Starting to process Excel files in: {directory_path}")
    process_directory(directory_path)
    print("Script completed successfully!")


if __name__ == "__main__":
    main()